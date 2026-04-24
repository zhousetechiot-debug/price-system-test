from __future__ import annotations

from datetime import datetime
from typing import Dict, Any
from pathlib import Path
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app import models
from app.services.pricing import (
    calculate_final_cost_twd,
    build_price_levels,
    safe_number,
    normalize_percentage_value,
    DEFAULT_DISCOUNT_RATIOS,
    PRICE_LEVELS,
    get_discount_price,
    calculate_duty_cost_usd,
)


def _clean_str(value):
    if value is None:
        return ""
    return str(value).strip()


def _upsert_rate(db: Session, currency: str, rate: float):
    row = db.query(models.ExchangeRate).filter(models.ExchangeRate.currency == currency).first()
    if row:
        row.rate_to_twd = rate
        row.updated_at = datetime.utcnow()
    else:
        row = models.ExchangeRate(currency=currency, rate_to_twd=rate)
        db.add(row)
    db.commit()



def ensure_default_rates(db: Session, usd_rate: float = 32.5, rmb_rate: float = 4.5):
    _upsert_rate(db, "USD", usd_rate)
    _upsert_rate(db, "RMB", rmb_rate)
    _upsert_rate(db, "TWD", 1.0)



def ensure_default_price_settings(db: Session):
    for level in PRICE_LEVELS:
        row = db.query(models.PriceSetting).filter(models.PriceSetting.level_name == level).first()
        if not row:
            db.add(models.PriceSetting(level_name=level, discount_ratio=DEFAULT_DISCOUNT_RATIOS[level]))
    db.commit()



def get_price_setting_map(db: Session) -> Dict[str, float]:
    ensure_default_price_settings(db)
    rows = db.query(models.PriceSetting).all()
    return {r.level_name: r.discount_ratio for r in rows}



def get_rate_map(db: Session) -> Dict[str, float]:
    rows = db.query(models.ExchangeRate).all()
    data = {r.currency: r.rate_to_twd for r in rows}
    data.setdefault("USD", 32.5)
    data.setdefault("RMB", 4.5)
    data.setdefault("TWD", 1.0)
    return data



def apply_discount_settings_to_products(db: Session):
    discount_map = get_price_setting_map(db)
    products = db.query(models.Product).all()
    for product in products:
        market_price = safe_number(product.market_price)
        if market_price <= 0 and product.final_cost_twd > 0:
            market_price = build_price_levels(product.final_cost_twd)["market_price"]
            product.market_price = market_price
        product.market_min_price = get_discount_price(market_price, discount_map.get("市場最低價", 95), getattr(product, 'special_market_min_ratio', 1))
        product.designer_price = get_discount_price(market_price, discount_map.get("設計師價", 90), getattr(product, 'special_designer_ratio', 1))
        product.dealer_lv1_price = get_discount_price(market_price, discount_map.get("一級經銷商", 85), getattr(product, 'special_dealer_lv1_ratio', 1))
        product.dealer_lv2_price = get_discount_price(market_price, discount_map.get("二級經銷商", 80), getattr(product, 'special_dealer_lv2_ratio', 1))
        product.branch_price = get_discount_price(market_price, discount_map.get("分公司", 75), getattr(product, 'special_branch_ratio', 1))
        product.master_dealer_price = get_discount_price(market_price, discount_map.get("總經銷商", 70), getattr(product, 'special_master_ratio', 1))
        product.updated_at = datetime.utcnow()
    db.commit()



def _normalize_ratio_value(value, allow_zero: bool = False) -> float:
    num = safe_number(value)
    if num == 0 and allow_zero:
        return 0.0
    if num == 0:
        return 1.0
    return round(num * 100.0, 3) / 100.0 if 0 < num <= 1 else round(num, 3)


def _normalize_pct_value(value) -> float:
    return normalize_percentage_value(value)


def _track_import_change(db: Session, product, before: dict, after: dict):
    labels = {
        'category': '類別', 'status': '狀態', 'name': '品名', 'description': '產品描述', 'unit': '單位',
        'source_currency': '來源幣別', 'source_cost': '來源成本', 'shipping_usd': '運費(USD)',
        'duty_rate_pct': '關稅/其他成本加成(%)', 'duty_cost_usd': '加乘後金額(USD)',
        'outsourced_parts_fee_twd': '其他加購(TWD)', 'final_cost_twd': '成本(TWD)',
        'planning_fee_twd': '規劃費', 'setup_fee_twd': '設定費',
        'special_market_min_ratio': '特殊倍數_市場最低價', 'special_designer_ratio': '特殊倍數_設計師價',
        'special_dealer_lv1_ratio': '特殊倍數_一級經銷商', 'special_dealer_lv2_ratio': '特殊倍數_二級經銷商',
        'special_branch_ratio': '特殊倍數_分公司', 'special_master_ratio': '特殊倍數_總經銷商',
        'market_price': '市場報價', 'market_min_price': '市場最低價', 'designer_price': '設計師價',
        'dealer_lv1_price': '一級經銷商', 'dealer_lv2_price': '二級經銷商', 'branch_price': '分公司',
        'master_dealer_price': '總經銷商', 'note': '備註',
    }
    changes = []
    for key, label in labels.items():
        bv = before.get(key)
        av = after.get(key)
        if str(bv if bv is not None else '') != str(av if av is not None else ''):
            changes.append(f'{label}: {bv if bv is not None else ""} → {av if av is not None else ""}')
    if changes:
        db.add(models.ProductChangeLog(product_id=product.id, changed_by='import', detail='\n'.join(changes)[:4000]))


def import_dealers(db: Session, workbook_path: str) -> int:
    wb = load_workbook(workbook_path, data_only=True)
    if "經銷商資料" not in wb.sheetnames:
        return 0

    ws = wb["經銷商資料"]
    count = 0
    for row in range(3, ws.max_row + 1):
        level = _clean_str(ws[f"B{row}"].value)
        name = _clean_str(ws[f"C{row}"].value)
        if not name:
            continue

        dealer = db.query(models.Dealer).filter(models.Dealer.name == name).first()
        if not dealer:
            dealer = models.Dealer(name=name)
            db.add(dealer)

        dealer.level = level
        dealer.tax_id = _clean_str(ws[f"D{row}"].value)
        dealer.address = _clean_str(ws[f"E{row}"].value)
        dealer.phone = _clean_str(ws[f"F{row}"].value)
        dealer.shipping_note = _clean_str(ws[f"G{row}"].value)
        dealer.order_note = _clean_str(ws[f"H{row}"].value)
        dealer.closing_day = _clean_str(ws[f"I{row}"].value)
        dealer.payment_day = _clean_str(ws[f"J{row}"].value)
        dealer.closing_note = _clean_str(ws[f"K{row}"].value)
        count += 1

    db.commit()
    return count



def _read_orvibo_map(workbook_path: str) -> Dict[str, Dict[str, Any]]:
    wb = load_workbook(workbook_path, data_only=True)
    result = {}
    if "ORVIBO" not in wb.sheetnames:
        return result

    ws = wb["ORVIBO"]
    for row in range(4, ws.max_row + 1):
        model = _clean_str(ws[f"D{row}"].value)
        if not model:
            continue
        result[model] = {
            "description_en": _clean_str(ws[f"E{row}"].value),
            "msrp_rmb": safe_number(ws[f"O{row}"].value),
            "dist_rmb": safe_number(ws[f"P{row}"].value),
            "msrp_usd": safe_number(ws[f"Q{row}"].value),
            "dist_usd": safe_number(ws[f"R{row}"].value),
        }
    return result





def _guess_image_suffix(image) -> str:
    fmt = (getattr(image, "format", "") or "").lower()
    if fmt in {"png", "jpeg", "jpg"}:
        return ".jpg" if fmt == "jpeg" else f".{fmt}"
    path = getattr(image, "path", "") or ""
    suffix = Path(path).suffix.lower()
    return suffix or ".png"


def _extract_row_image_map(workbook_path: str, sheet_name: str = "智慧家庭產品價格表.含稅") -> Dict[int, Dict[str, Any]]:
    wb = load_workbook(workbook_path)
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    result: Dict[int, Dict[str, Any]] = {}
    for image in getattr(ws, "_images", []):
        try:
            anchor = getattr(image, "anchor", None)
            anchor_from = getattr(anchor, "_from", None)
            if anchor_from is None:
                continue
            row_no = int(anchor_from.row) + 1
            col_no = int(anchor_from.col)
            if col_no != 9:  # J column in 0-indexed anchor coords
                continue
            data = None
            ref = getattr(image, "ref", None)
            if hasattr(ref, "getvalue"):
                data = ref.getvalue()
            elif hasattr(image, "_data"):
                data = image._data()
            if data:
                result[row_no] = {
                    "bytes": data,
                    "suffix": _guess_image_suffix(image),
                }
        except Exception:
            continue
    return result


def _write_product_image(workbook_path: str, model: str, row_no: int, image_bytes: bytes, image_suffix: str = ".png") -> str:
    workbook_dir = Path(workbook_path).resolve().parent
    image_dir = workbook_dir / "uploads" / "images"
    image_dir.mkdir(parents=True, exist_ok=True)
    safe_model = "".join(ch if ch.isalnum() or ch in {"-", "_"} else "_" for ch in (model or f"row_{row_no}"))
    file_name = f"{safe_model}{image_suffix}"
    file_path = image_dir / file_name
    if not file_path.exists() or file_path.stat().st_size == 0:
        file_path.write_bytes(image_bytes)
    return f"/uploads/images/{file_name}"


def import_products(db: Session, workbook_path: str) -> int:
    wb = load_workbook(workbook_path, data_only=True)
    if "智慧家庭產品價格表.含稅" not in wb.sheetnames:
        return 0

    rates = get_rate_map(db)
    orvibo_map = _read_orvibo_map(workbook_path)
    ws = wb["智慧家庭產品價格表.含稅"]
    image_bytes_map = _extract_row_image_map(workbook_path, "智慧家庭產品價格表.含稅")

    count = 0
    product_cache = {p.model: p for p in db.query(models.Product).all()}
    for row in range(6, ws.max_row + 1):
        category = _clean_str(ws[f"B{row}"].value)
        model = _clean_str(ws[f"C{row}"].value)
        name = _clean_str(ws[f"D{row}"].value)

        if not model or not name:
            continue

        duty_cost_twd = safe_number(ws[f"O{row}"].value)
        shipping_usd = 0.0
        duty_rate_pct = 0.0

        source_cost = 0.0
        source_currency = "TWD"
        if model in orvibo_map:
            ov = orvibo_map[model]
            if safe_number(ov.get("dist_usd")) > 0:
                source_cost = safe_number(ov.get("dist_usd"))
                source_currency = "USD"
            elif safe_number(ov.get("dist_rmb")) > 0:
                source_cost = safe_number(ov.get("dist_rmb"))
                source_currency = "RMB"
        if source_cost <= 0:
            source_cost = safe_number(ws[f"N{row}"].value) or safe_number(ws[f"L{row}"].value)
            source_currency = "USD" if source_cost > 0 else "TWD"

        final_cost_twd = calculate_final_cost_twd(
            source_currency=source_currency,
            source_cost=source_cost,
            usd_rate=rates["USD"],
            rmb_rate=rates["RMB"],
            duty_cost_twd=duty_cost_twd,
        )
        prices = build_price_levels(final_cost_twd)

        product = product_cache.get(model)
        if not product:
            product = models.Product(model=model, name=name)
            db.add(product)
            product_cache[model] = product

        product.category = category
        product.name = name
        product.description = _clean_str(ws[f"H{row}"].value) or orvibo_map.get(model, {}).get("description_en", "")
        product.note = _clean_str(ws[f"I{row}"].value)
        image_cell_value = _clean_str(ws[f"J{row}"].value)
        if image_cell_value:
            normalized_image_value = image_cell_value.replace('\\', '/').strip()
            if normalized_image_value.startswith('http://') or normalized_image_value.startswith('https://'):
                product.image_url = normalized_image_value
                product.image_path = None
            else:
                if normalized_image_value.startswith('/uploads/'):
                    normalized_image_value = normalized_image_value[len('/uploads/'): ]
                normalized_image_value = normalized_image_value.lstrip('/')
                product.image_path = normalized_image_value
                product.image_url = None
        elif row in image_bytes_map:
            image_info = image_bytes_map[row]
            product.image_path = _write_product_image(
                workbook_path,
                model,
                row,
                image_info.get('bytes', b''),
                image_info.get('suffix', '.png'),
            )
            product.image_url = None
        product.unit = _clean_str(ws[f"K{row}"].value) or '台'
        product.status = _clean_str(ws[f"AI{row}"].value)
        product.special_ratio = safe_number(ws[f"AJ{row}"].value)

        product.source_currency = source_currency
        product.source_cost = source_cost
        product.shipping_usd = shipping_usd
        product.duty_rate_pct = duty_rate_pct
        product.duty_cost_usd = calculate_duty_cost_usd(source_currency, source_cost, shipping_usd, duty_rate_pct, rates['USD'], rates['RMB'])
        product.duty_cost_twd = duty_cost_twd
        product.final_cost_twd = final_cost_twd
        product.planning_fee_twd = safe_number(ws[f"AA{row}"].value)
        product.setup_fee_twd = safe_number(ws[f"AB{row}"].value)
        product.special_market_min_ratio = 1.0
        product.special_designer_ratio = 1.0
        product.special_dealer_lv1_ratio = 1.0
        product.special_dealer_lv2_ratio = 1.0
        product.special_branch_ratio = 1.0
        product.special_master_ratio = 1.0

        product.market_price = round(safe_number(ws[f"AD{row}"].value) or prices["market_price"],0)
        product.market_min_price = round(safe_number(ws[f"BK{row}"].value) or prices["market_min_price"],0)
        product.designer_price = round(safe_number(ws[f"BF{row}"].value) or prices["designer_price"],0)
        product.dealer_lv1_price = round(safe_number(ws[f"AV{row}"].value) or prices["dealer_lv1_price"],0)
        product.dealer_lv2_price = round(safe_number(ws[f"BA{row}"].value) or prices["dealer_lv2_price"],0)
        product.branch_price = round(safe_number(ws[f"AM{row}"].value) or prices["branch_price"],0)
        product.master_dealer_price = round(safe_number(ws[f"AQ{row}"].value) or prices["master_dealer_price"],0)
        product.updated_at = datetime.utcnow()

        count += 1

    db.commit()
    return count


def import_all(db: Session, workbook_path: str, usd_rate: float | None = None, rmb_rate: float | None = None) -> dict:
    current = get_rate_map(db)
    ensure_default_rates(
        db,
        usd_rate=usd_rate if usd_rate is not None else current.get("USD", 32.5),
        rmb_rate=rmb_rate if rmb_rate is not None else current.get("RMB", 4.5),
    )
    ensure_default_price_settings(db)

    dealer_count = import_dealers(db, workbook_path)
    product_count = import_system_sheet_products(db, workbook_path)
    if product_count == 0:
        product_count = import_products(db, workbook_path)

    return {
        "dealers": dealer_count,
        "products": product_count,
    }


def _find_header_row(rows: list[tuple], required_headers: list[str], scan_limit: int = 10) -> int | None:
    for row_index, row in enumerate(rows[:scan_limit]):
        normalized = {_clean_str(cell) for cell in row if _clean_str(cell)}
        if all(header in normalized for header in required_headers):
            return row_index
    return None



def import_system_sheet_products(db: Session, workbook_path: str) -> int:
    wb_formula = load_workbook(workbook_path, data_only=False)
    if '系統匯入資料' not in wb_formula.sheetnames:
        return 0

    # 先讀公式版，保留可能尚未被 Excel 寫回快取值的內容；
    # 另外再讀一份 data_only 版，若使用者先前有在 Excel 存檔，就能優先取到實際計算值。
    wb_values = load_workbook(workbook_path, data_only=True)
    ws_formula = wb_formula['系統匯入資料']
    ws_values = wb_values['系統匯入資料'] if '系統匯入資料' in wb_values.sheetnames else None

    rows_formula = list(ws_formula.iter_rows(values_only=True))
    rows_values = list(ws_values.iter_rows(values_only=True)) if ws_values else []
    if not rows_formula:
        return 0

    header_row_index = _find_header_row(rows_formula, ['類別', '狀態', '型號', '品名'])
    if header_row_index is None:
        return 0

    headers = [_clean_str(h) for h in rows_formula[header_row_index]]
    idx = {h: i for i, h in enumerate(headers) if h}

    alias_map = {
        '來源成本': ['來源成本', '來源成本(USD)'],
        '關稅/其他成本加成(%)': ['關稅/其他成本加成(%)'],
        '其他加購(TWD)': ['其他加購(TWD)', '外採配件費用(TWD)'],
        '市場報價': ['市場報價', '市場價'],
        '特殊倍數_市場價': ['特殊倍數_市場價'],
        '市場最低價': ['市場最低價'],
        '設計師價': ['設計師價'],
        '一級經銷商': ['一級經銷商'],
        '二級經銷商': ['二級經銷商'],
        '分公司': ['分公司'],
        '總經銷商': ['總經銷商'],
    }

    def _get_from_row(row, key, default=''):
        possible_keys = alias_map.get(key, [key])
        for candidate in possible_keys:
            col_index = idx.get(candidate)
            if col_index is not None and col_index < len(row):
                value = row[col_index]
                if value is not None:
                    return value
        return default

    def val(row_formula, row_values, key, default=''):
        value_from_values = _get_from_row(row_values, key, None) if row_values else None
        if value_from_values not in (None, ''):
            return value_from_values
        value_from_formula = _get_from_row(row_formula, key, None)
        if value_from_formula not in (None, ''):
            return value_from_formula
        return default

    product_cache = {p.model: p for p in db.query(models.Product).all()}
    count = 0
    for data_row_index in range(header_row_index + 1, len(rows_formula)):
        row_formula = rows_formula[data_row_index]
        row_values = rows_values[data_row_index] if data_row_index < len(rows_values) else None

        model = _clean_str(val(row_formula, row_values, '型號'))
        name = _clean_str(val(row_formula, row_values, '品名'))
        if not model or not name:
            continue

        product = product_cache.get(model)
        if not product:
            product = models.Product(model=model, name=name or model)
            db.add(product)
            product_cache[model] = product

        product.category = _clean_str(val(row_formula, row_values, '類別'))
        product.status = _clean_str(val(row_formula, row_values, '狀態'))
        product.name = name or model
        product.unit = _clean_str(val(row_formula, row_values, '單位')) or '台'
        product.description = _clean_str(val(row_formula, row_values, '產品描述'))

        img = _clean_str(val(row_formula, row_values, '圖片'))
        if img:
            normalized_img = img.replace('\\', '/').strip()
            if normalized_img.startswith('http://') or normalized_img.startswith('https://'):
                product.image_url = normalized_img
                product.image_path = None
            else:
                product.image_path = normalized_img.replace('/uploads/', '').lstrip('/')
                product.image_url = None

        product.source_currency = _clean_str(val(row_formula, row_values, '來源幣別')) or 'TWD'
        before = {k: getattr(product, k, None) for k in ['category','status','name','description','unit','source_currency','source_cost','shipping_usd','duty_rate_pct','duty_cost_usd','outsourced_parts_fee_twd','final_cost_twd','planning_fee_twd','setup_fee_twd','special_market_min_ratio','special_designer_ratio','special_dealer_lv1_ratio','special_dealer_lv2_ratio','special_branch_ratio','special_master_ratio','market_price','market_min_price','designer_price','dealer_lv1_price','dealer_lv2_price','branch_price','master_dealer_price','note']}

        product.source_cost = safe_number(val(row_formula, row_values, '來源成本'))
        product.shipping_usd = safe_number(val(row_formula, row_values, '運費(USD)'))
        product.duty_rate_pct = _normalize_pct_value(val(row_formula, row_values, '關稅/其他成本加成(%)'))
        product.duty_cost_usd = safe_number(val(row_formula, row_values, '加乘後金額(USD)'))
        product.outsourced_parts_fee_twd = safe_number(val(row_formula, row_values, '其他加購(TWD)'))
        product.final_cost_twd = safe_number(val(row_formula, row_values, '成本(TWD)'))
        product.planning_fee_twd = safe_number(val(row_formula, row_values, '規劃費'))
        product.setup_fee_twd = safe_number(val(row_formula, row_values, '設定費'))
        product.special_market_min_ratio = _normalize_ratio_value(val(row_formula, row_values, '特殊倍數_市場最低價'))
        product.special_designer_ratio = _normalize_ratio_value(val(row_formula, row_values, '特殊倍數_設計師價'))
        product.special_dealer_lv1_ratio = _normalize_ratio_value(val(row_formula, row_values, '特殊倍數_一級經銷商'))
        product.special_dealer_lv2_ratio = _normalize_ratio_value(val(row_formula, row_values, '特殊倍數_二級經銷商'))
        product.special_branch_ratio = _normalize_ratio_value(val(row_formula, row_values, '特殊倍數_分公司'))
        product.special_master_ratio = _normalize_ratio_value(val(row_formula, row_values, '特殊倍數_總經銷商'))
        product.market_price = round(safe_number(val(row_formula, row_values, '市場報價')), 0)
        product.market_min_price = round(safe_number(val(row_formula, row_values, '市場最低價')), 0)
        product.designer_price = round(safe_number(val(row_formula, row_values, '設計師價')), 0)
        product.dealer_lv1_price = round(safe_number(val(row_formula, row_values, '一級經銷商')), 0)
        product.dealer_lv2_price = round(safe_number(val(row_formula, row_values, '二級經銷商')), 0)
        product.branch_price = round(safe_number(val(row_formula, row_values, '分公司')), 0)
        product.master_dealer_price = round(safe_number(val(row_formula, row_values, '總經銷商')), 0)
        product.note = _clean_str(val(row_formula, row_values, '備註'))
        product.updated_at = datetime.utcnow()
        db.flush()
        after = {k: getattr(product, k, None) for k in before.keys()}
        _track_import_change(db, product, before, after)
        count += 1

    db.commit()
    return count
