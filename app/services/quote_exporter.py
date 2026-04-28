from __future__ import annotations

from pathlib import Path
import json
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.pagebreak import Break
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, KeepTogether

pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
FONT_NAME = "STSong-Light"

DEFAULT_NOTE = """1.報價支付方式
   總價低於15萬，訂金60% 尾款40%。
   總價高於15萬，訂金30% 設備進場40% 尾款30%。
2.若開立支票請開立即期票。
3.報價內容不含水電工程。(例:開關、燈具安裝)
4.自驗收發票日起算保固1年(非人為損壞)。
5.報價有效期自報價日起14日曆天止。
6.客製化設備交期為8週。
7.代購商品需全額支付，恕不退換。"""


def _local_image_path(base_dir: str, image_ref: str | None):
    if not image_ref:
        return None
    ref = str(image_ref or '').strip()
    base = Path(base_dir or '')
    if ref.startswith('/uploads/'):
        p = base / ref.lstrip('/')
        if p.exists():
            return p
        p2 = base.parent / ref.lstrip('/')
        return p2 if p2.exists() else None
    p = Path(ref)
    if p.exists():
        return p
    p = base / ref.lstrip('/')
    return p if p.exists() else None


def _company_logo(base_dir: str | None):
    base = Path(base_dir or '')
    candidates = [
        base / 'uploads/company_logo.png',
        base / 'uploads/company_logo.jpg',
        base / 'uploads/company_logo.jpeg',
        base.parent / 'uploads/company_logo.png',
        base.parent / 'uploads/company_logo.jpg',
        base.parent / 'uploads/company_logo.jpeg',
    ]
    for p in candidates:
        if p.exists():
            return p
    return None


def _num(value):
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def _money(value):
    return int(round(_num(value), 0))


def build_display_items(quote):
    items = list(getattr(quote, 'items', []) or [])
    extra = []

    def obj(**kwargs):
        return type('Obj', (), kwargs)()

    try:
        curtain_rows = json.loads(getattr(quote, 'curtain_rows_json', '') or '[]')
    except Exception:
        curtain_rows = []
    for row in curtain_rows:
        qty = int(round(_num(row.get('qty')), 0))
        unit_price = _money(row.get('unit_price'))
        line_total = _money(row.get('line_total'))
        if qty and unit_price:
            space = row.get('space') or ''
            ctype = row.get('type') or '窗簾'
            track_length = int(_num(row.get('track_length')))
            cloth_height = int(_num(row.get('cloth_height') or row.get('height')))
            pname = f"{space}-{ctype} {track_length}cm×{cloth_height}cm".strip('- ')
            extra.append(obj(image_path='', model='', product_name=pname, qty=qty, unit='組', unit_price_twd=unit_price, line_total_twd=line_total, note=(row.get('note') or getattr(quote, 'curtain_note', '') or '')))

    if _money(getattr(quote, 'planning_fee_total', 0)):
        total = _money(getattr(quote, 'planning_fee_total', 0))
        extra.append(obj(image_path='', model='', product_name='規劃費', qty=1, unit='式', unit_price_twd=total, line_total_twd=total, note=''))
    if _money(getattr(quote, 'setup_fee_total', 0)):
        total = _money(getattr(quote, 'setup_fee_total', 0))
        extra.append(obj(image_path='', model='', product_name='設定費', qty=1, unit='式', unit_price_twd=total, line_total_twd=total, note=''))
    if _money(getattr(quote, 'dispatch_fee', 0)):
        total = _money(getattr(quote, 'dispatch_fee', 0))
        extra.append(obj(image_path='', model='', product_name='派工費', qty=1, unit='式', unit_price_twd=total, line_total_twd=total, note=getattr(quote, 'dispatch_label', '') or ''))
    if _money(getattr(quote, 'lock_install_fee', 0)):
        qty = int(_num(getattr(quote, 'lock_install_qty', 0))) or 1
        unit_price = _money(getattr(quote, 'lock_install_unit_price', 0))
        total = _money(getattr(quote, 'lock_install_fee', 0))
        extra.append(obj(image_path='', model='', product_name='門鎖施工費', qty=qty, unit='把', unit_price_twd=unit_price, line_total_twd=total, note=''))
    if _money(getattr(quote, 'curtain_install_amount', 0)) and _money(getattr(quote, 'curtain_install_qty', 0)):
        qty = _money(getattr(quote, 'curtain_install_qty', 0))
        unit_price = _money(getattr(quote, 'curtain_install_amount', 0))
        extra.append(obj(image_path='', model='', product_name='窗簾施工', qty=qty, unit=getattr(quote, 'curtain_install_unit', '組') or '組', unit_price_twd=unit_price, line_total_twd=_money(unit_price * qty), note='窗簾施工'))
    for key, label in [('weak_current', '弱電費用'), ('hardware', '五金材料'), ('water_elec', '水電費用')]:
        amount = _money(getattr(quote, f'{key}_amount', 0))
        if amount:
            qty = _money(getattr(quote, f'{key}_qty', 1)) or 1
            unit = getattr(quote, f'{key}_unit', '式') or '式'
            default_note = {'weak_current': '弱電材料與施工整體費用', 'hardware': '螺絲、固定片、耗材等', 'water_elec': '外部水電配合施工費用'}.get(key, '')
            extra.append(obj(image_path='', model='', product_name=label, qty=qty, unit=unit, unit_price_twd=_money(amount / qty) if qty else amount, line_total_twd=amount, note=default_note))
    try:
        custom_rows = json.loads(getattr(quote, 'custom_fee_json', '') or '[]')
    except Exception:
        custom_rows = []
    for row in custom_rows:
        total = _money(row.get('total'))
        if total:
            extra.append(obj(image_path='', model='', product_name=row.get('name') or '自訂工費', qty=_money(row.get('qty')) or 1, unit=row.get('unit') or '式', unit_price_twd=_money(row.get('unit_price')), line_total_twd=total, note=row.get('note') or ''))
    return items + extra


def export_quote_to_excel(quote, output_path: str, base_dir: str | None = None):
    wb = Workbook()
    ws = wb.active
    ws.title = '報價單'

    # v84/v86：Excel 改為 A4 直式，寬度縮成 1 頁，且金額欄保留公式。
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.print_options.horizontalCentered = True

    widths = {'A': 8, 'B': 16, 'C': 26, 'D': 9, 'E': 8, 'F': 12, 'G': 13, 'H': 22}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    thin = Side(style='thin', color='999999')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    fill = PatternFill('solid', fgColor='EAF2FF')

    ws.merge_cells('C1:F1')
    ws['C1'] = '報價單'
    ws['C1'].font = Font(size=18, bold=True)
    ws['C1'].alignment = center
    logo = _company_logo(base_dir)
    if logo:
        try:
            img = XLImage(str(logo))
            img.width = 115
            img.height = 42
            ws.add_image(img, 'A1')
        except Exception:
            pass

    meta = [
        (3, '訂購單號', getattr(quote, 'quote_no', ''), 'From', '智崴物聯科技股份有限公司'),
        (4, 'Name :', getattr(quote, 'customer_name', ''), 'Date :', str(getattr(quote, 'quote_date', ''))),
        # v87 fix: 業務姓名改帶到 G5/H5（ATTN），不要放在 A9/B9。
        (5, 'Contact :', getattr(quote, 'contact_name', '') or '', 'ATTN :', getattr(quote, 'sales_name', '') or getattr(quote, 'attn', '') or ''),
        (6, 'Tel :', getattr(quote, 'phone', '') or '', 'TEL :', getattr(quote, 'sales_phone', '') or ''),
        (7, 'Mail :', getattr(quote, 'email', '') or '', 'Mail :', getattr(quote, 'sales_email', '') or ''),
        (8, 'Add :', getattr(quote, 'address', '') or '', 'TEL(O) :', '(02) 2600-7081'),
        # v87 fix: A9/B9 需為空白，右側保留幣別。
        (9, '', '', '幣別', 'NTD'),
    ]
    for row, l1, v1, l2, v2 in meta:
        ws[f'A{row}'] = l1
        ws.merge_cells(f'B{row}:D{row}')
        ws[f'B{row}'] = v1
        ws[f'F{row}'] = l2
        ws.merge_cells(f'G{row}:H{row}')
        ws[f'G{row}'] = v2
        for cell in [f'A{row}', f'B{row}', f'F{row}', f'G{row}']:
            ws[cell].alignment = left
            ws[cell].border = border

    headers = ['項目', '型號', '品名', '數量', '單位', '單價', '總價', '備註']
    start_row = 11
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=i, value=h)
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.border = border
        cell.fill = fill

    display_items = build_display_items(quote)
    total_rows = max(15, len(display_items))
    first_item_row = start_row + 1
    row = first_item_row
    items_per_page = 15
    for idx in range(total_rows):
        item = display_items[idx] if idx < len(display_items) else None
        ws.row_dimensions[row].height = 34
        qty = int(round(_num(getattr(item, 'qty', 0)), 0)) if item else ''
        unit_price = _money(getattr(item, 'unit_price_twd', 0)) if item else ''
        vals = [
            idx + 1 if item else '',
            getattr(item, 'model', '') if item else '',
            getattr(item, 'product_name', '') if item else '',
            qty,
            getattr(item, 'unit', '') if item else '',
            unit_price,
            '',
            getattr(item, 'note', '') if item else '',
        ]
        for col, value in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = center if col not in (3, 8) else left
        if item:
            # 公式：總價 = 數量 × 單價；讓輸出後 Excel 可自行重新計算。
            ws[f'G{row}'] = f'=D{row}*F{row}'
            ws[f'F{row}'].number_format = '#,##0'
            ws[f'G{row}'].number_format = '#,##0'
        if idx > 0 and idx % items_per_page == 0:
            ws.row_breaks.append(Break(id=row - 1))
        row += 1

    subtotal_row = row + 1
    tax_row = subtotal_row + 1
    total_row = tax_row + 1
    final_row = total_row

    summary_rows = [
        (subtotal_row, '未稅小計', f'=SUM(G{first_item_row}:G{row - 1})'),
        (tax_row, '稅額', f'=ROUND(G{subtotal_row}*0.05,0)'),
        (total_row, '合計', f'=G{subtotal_row}+G{tax_row}'),
    ]

    # v87 fix: 訂金 / 設備進場 / 驗收款改成 Excel 公式，而不是只輸出固定數值。
    payment_base_row = total_row
    if _money(getattr(quote, 'negotiated_total', 0)):
        final_row += 1
        negotiated_row = final_row
        summary_rows.append((negotiated_row, '議價後合計', _money(getattr(quote, 'negotiated_total', 0))))
        payment_base_row = negotiated_row

    payment_scheme = (getattr(quote, 'payment_scheme', '') or '').strip()
    payment_base = f'G{payment_base_row}'
    if payment_scheme == '30/40/30':
        final_row += 1
        summary_rows.append((final_row, '訂金', f'=ROUND({payment_base}*30%,0)'))
        final_row += 1
        summary_rows.append((final_row, '設備進場', f'=ROUND({payment_base}*40%,0)'))
        final_row += 1
        summary_rows.append((final_row, '尾款(驗收款)', f'=ROUND({payment_base}*30%,0)'))
    elif payment_scheme == '60/40':
        final_row += 1
        summary_rows.append((final_row, '訂金', f'=ROUND({payment_base}*60%,0)'))
        final_row += 1
        summary_rows.append((final_row, '尾款(驗收款)', f'=ROUND({payment_base}*40%,0)'))
    elif payment_scheme and payment_scheme != '一次付清':
        # 其他付款條件仍保留公式基礎，避免輸出成死數值。
        final_row += 1
        summary_rows.append((final_row, '訂金', f'=ROUND({payment_base}*60%,0)'))
        final_row += 1
        summary_rows.append((final_row, '尾款(驗收款)', f'=ROUND({payment_base}*40%,0)'))

    for r, label, value in summary_rows:
        ws[f'F{r}'] = label
        ws[f'G{r}'] = value
        ws.merge_cells(f'G{r}:H{r}')
        ws[f'F{r}'].border = border
        ws[f'G{r}'].border = border
        ws[f'F{r}'].alignment = left
        ws[f'G{r}'].alignment = left
        ws[f'G{r}'].number_format = '#,##0'

    note_start = final_row + 2
    note_text = '備註：\n' + (getattr(quote, 'note', '') or DEFAULT_NOTE)
    note_rows = max(10, len(note_text.splitlines()) + 2)
    note_end = note_start + note_rows - 1
    ws.merge_cells(start_row=note_start, start_column=1, end_row=note_end, end_column=4)
    note_cell = ws.cell(row=note_start, column=1, value=note_text)
    note_cell.alignment = Alignment(wrap_text=True, vertical='top')
    note_cell.font = Font(size=10)

    payment_rows = [
        ('付款方式:', '匯款'),
        ('戶名:', '智崴物聯科技股份有限公司'),
        ('銀行:', '上海商業儲蓄銀行-林口分行'),
        ('帳號:', '79102-0000-66086'),
    ]
    for idx, (label, value) in enumerate(payment_rows):
        r = note_start + idx
        ws[f'F{r}'] = label
        ws.merge_cells(f'G{r}:H{r}')
        ws[f'G{r}'] = value
        ws[f'F{r}'].alignment = left
        ws[f'G{r}'].alignment = left

    sign_row = max(note_end, note_start + len(payment_rows) + 1) + 1
    ws[f'A{sign_row}'] = '客戶回簽'
    ws.merge_cells(f'F{sign_row}:H{sign_row}')
    ws[f'F{sign_row}'] = '(回簽後, 視同訂單生效)'
    ws.print_title_rows = '1:11'
    ws.print_area = f'A1:H{sign_row}'
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _p(text, style):
    text = '' if text is None else str(text)
    return Paragraph(text.replace('\n', '<br/>'), style)


def export_quote_to_pdf(quote, output_path: str, base_dir: str | None = None):
    """A4 直式 PDF，自動縮欄、換行、分頁，避免內容超出頁面。"""
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    styles = getSampleStyleSheet()
    normal = ParagraphStyle('zh', parent=styles['Normal'], fontName=FONT_NAME, fontSize=7.2, leading=9, alignment=TA_LEFT)
    small = ParagraphStyle('zhSmall', parent=normal, fontSize=6.4, leading=8)
    title_style = ParagraphStyle('titleZh', parent=normal, fontSize=16, leading=20, alignment=TA_CENTER)
    header_style = ParagraphStyle('headerZh', parent=normal, fontSize=7.2, leading=8.5, alignment=TA_CENTER)

    story = []
    logo = _company_logo(base_dir)
    if logo:
        try:
            story.append(Table([[RLImage(str(logo), width=38*mm, height=14*mm), _p('報價單', title_style), '']], colWidths=[45*mm, 90*mm, 45*mm]))
        except Exception:
            story.append(_p('報價單', title_style))
    else:
        story.append(_p('報價單', title_style))
    story.append(Spacer(1, 3 * mm))

    meta_left = [
        ['訂購單號', getattr(quote, 'quote_no', '')],
        ['Name :', getattr(quote, 'customer_name', '')],
        ['Contact :', getattr(quote, 'contact_name', '') or ''],
        ['Tel :', getattr(quote, 'phone', '') or ''],
        ['Mail :', getattr(quote, 'email', '') or ''],
        ['Add :', getattr(quote, 'address', '') or ''],
        ['', ''],
    ]
    meta_right = [
        ['From', '智崴物聯科技股份有限公司'],
        ['Date :', str(getattr(quote, 'quote_date', ''))],
        ['ATTN :', getattr(quote, 'sales_name', '') or getattr(quote, 'attn', '') or ''],
        ['TEL :', getattr(quote, 'sales_phone', '') or ''],
        ['Mail :', getattr(quote, 'sales_email', '') or ''],
        ['TEL(O) :', '(02) 2600-7081'],
        ['幣別', 'NTD'],
    ]
    meta_rows = []
    for l, r in zip(meta_left, meta_right):
        meta_rows.append([_p(l[0], normal), _p(l[1], normal), _p(r[0], normal), _p(r[1], normal)])
    meta_table = Table(meta_rows, colWidths=[19*mm, 70*mm, 19*mm, 70*mm], repeatRows=0)
    meta_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.45, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 2),
        ('RIGHTPADDING', (0,0), (-1,-1), 2),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 4 * mm))

    headers = ['項目', '圖片', '型號', '品名', '數量', '單位', '單價', '總價', '備註']
    data = [[_p(h, header_style) for h in headers]]
    for idx, item in enumerate(build_display_items(quote), start=1):
        img_cell = ''
        local_img = _local_image_path(base_dir or '', getattr(item, 'image_path', ''))
        if local_img:
            try:
                img_cell = RLImage(str(local_img), width=12*mm, height=10*mm)
            except Exception:
                img_cell = ''
        data.append([
            _p(idx, small),
            img_cell,
            _p(getattr(item, 'model', '') or '', small),
            _p(getattr(item, 'product_name', '') or '', small),
            _p(int(round(_num(getattr(item, 'qty', 0)), 0)), small),
            _p(getattr(item, 'unit', '') or '', small),
            _p(f"{_money(getattr(item, 'unit_price_twd', 0)):,}", small),
            _p(f"{_money(getattr(item, 'line_total_twd', 0)):,}", small),
            _p(getattr(item, 'note', '') or '', small),
        ])
    min_rows = 12
    while len(data) < min_rows + 1:
        data.append([_p('', small) for _ in headers])

    col_widths = [10*mm, 15*mm, 24*mm, 42*mm, 10*mm, 10*mm, 16*mm, 17*mm, 34*mm]
    table = Table(data, colWidths=col_widths, repeatRows=1, splitByRow=1)
    table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.45, colors.black),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F1F5F9')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('LEFTPADDING', (0,0), (-1,-1), 1.5),
        ('RIGHTPADDING', (0,0), (-1,-1), 1.5),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]))
    story.append(table)
    story.append(Spacer(1, 4 * mm))

    summary_rows = [
        ['產品合計', f"{_money(getattr(quote, 'product_subtotal', 0)):,}"],
        ['規劃費', f"{_money(getattr(quote, 'planning_fee_total', 0)):,}"],
        ['設定費', f"{_money(getattr(quote, 'setup_fee_total', 0)):,}"],
        ['未稅小計', f"{_money(getattr(quote, 'subtotal', 0)):,}"],
        ['稅額', f"{_money(getattr(quote, 'tax_amount', 0)):,}"],
        ['合計', f"{_money(getattr(quote, 'total_amount', 0)):,}"],
    ]
    if _money(getattr(quote, 'negotiated_total', 0)):
        summary_rows.append(['議價後合計', f"{_money(getattr(quote, 'negotiated_total', 0)):,}"])
    if (getattr(quote, 'payment_scheme', '') or '') != '一次付清':
        summary_rows.append(['訂金', f"{_money(getattr(quote, 'deposit_1', 0)):,}"])
    if _money(getattr(quote, 'deposit_2', 0)):
        summary_rows.append(['設備進場', f"{_money(getattr(quote, 'deposit_2', 0)):,}"])
    if _money(getattr(quote, 'deposit_3', 0)) and (getattr(quote, 'payment_scheme', '') or '') != '一次付清':
        summary_rows.append(['尾款(驗收款)', f"{_money(getattr(quote, 'deposit_3', 0)):,}"])

    note = _p('備註：<br/>' + (getattr(quote, 'note', '') or DEFAULT_NOTE).replace('\n', '<br/>'), small)
    summary_table = Table([[note, Table([[_p(k, small), _p(v, small)] for k, v in summary_rows], colWidths=[24*mm, 31*mm])]], colWidths=[118*mm, 58*mm])
    summary_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.45, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 3),
        ('RIGHTPADDING', (0,0), (-1,-1), 3),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 3 * mm))

    bank_rows = [
        [_p('付款方式：匯款', small), _p('戶名：智崴物聯科技股份有限公司', small)],
        [_p('銀行：上海商業儲蓄銀行-林口分行', small), _p('帳號：79102-0000-66086', small)],
        [_p('客戶回簽', normal), _p('(回簽後，視同訂單生效)', small)],
    ]
    bank_table = Table(bank_rows, colWidths=[88*mm, 88*mm])
    bank_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
    story.append(bank_table)

    doc.build(story)
    return output_path
