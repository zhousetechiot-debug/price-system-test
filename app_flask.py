from __future__ import annotations

import json
import re
import shutil
import os
import math
from datetime import date, datetime
from pathlib import Path
from io import BytesIO
from PIL import Image, ImageOps

from flask import Flask, render_template, request, redirect, send_file, send_from_directory, abort, session, url_for, flash
from sqlalchemy import or_, text
from functools import wraps
from openpyxl import Workbook
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from app.database import Base, engine, SessionLocal
from app import models
from app.services.excel_importer import (
    ensure_default_rates,
    ensure_default_price_settings,
    get_rate_map,
    import_all,
    get_price_setting_map,
    apply_discount_settings_to_products,
)
from app.services.pricing import (
    get_price_by_level,
    get_profit_rate,
    calculate_final_cost_twd,
    build_price_levels,
    PRICE_LEVELS,
    safe_number,
    calculate_duty_cost_usd,
    round_int,
    normalize_percentage_value,
    DEFAULT_DISCOUNT_RATIOS,
)
from app.services.quote_exporter import export_quote_to_excel, export_quote_to_pdf, DEFAULT_NOTE

APP_NAME = "智崴物聯價格系統"
BASE_DIR = Path(__file__).resolve().parent
APP_DIR = BASE_DIR / "app"
UPLOAD_DIR = BASE_DIR / "uploads"
EXPORT_DIR = BASE_DIR / "exports"
IMAGE_DIR = UPLOAD_DIR / "images"
ORIGINAL_IMAGE_DIR = UPLOAD_DIR / "image_originals"
DB_PATH = BASE_DIR / "price_system.db"

UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
EXPORT_DIR.mkdir(parents=True, exist_ok=True)
IMAGE_DIR.mkdir(parents=True, exist_ok=True)
ORIGINAL_IMAGE_DIR.mkdir(parents=True, exist_ok=True)

flask_app = Flask(
    __name__,
    template_folder=str(APP_DIR / "templates"),
    static_folder=str(APP_DIR / "static"),
    static_url_path="/static",
)
flask_app.secret_key = 'zhiwei-v48-secret-key'
DEFAULT_INTERNAL_PASSWORD = 'zz55625107'
DEFAULT_SUPERVISOR_USERNAME = 'admin'
DEFAULT_SUPERVISOR_DISPLAY_NAME = '系統管理員'

Base.metadata.create_all(bind=engine)

DISPATCH_OPTIONS = [
    ("不派工", 0),
    ("基本派工 800", 800),
    ("基本派工 1200", 1200),
    ("基本派工 1500", 1500),
    ("半天 2000", 2000),
    ("一天 3500", 3500),
    ("一天 4500", 4500),
]



def safe_int(value, default=0):
    try:
        if value is None:
            return default
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, (int, float)):
            return int(round(float(value)))
        s = str(value).strip().replace(',', '')
        if s == '':
            return default
        return int(round(float(s)))
    except Exception:
        return default


def round3(value, default=1.0):
    try:
        if value is None:
            return default
        s = str(value).strip().replace(',', '')
        if s == '':
            return default
        return round(float(s), 3)
    except Exception:
        return default


INTERNAL_ROLE_PERMISSIONS = {
    'supervisor': {'all': True},
    'quote_only': {'quote_only': True},
}


def current_internal_user(db=None):
    user_id = session.get('internal_user_id')
    if not user_id:
        return None
    close_db = False
    if db is None:
        db = db_session()
        close_db = True
    try:
        return db.query(models.InternalUser).filter(models.InternalUser.id == user_id).first()
    finally:
        if close_db:
            db.close()


def is_internal_logged_in():
    return bool(session.get('internal_user_id') or session.get('internal_auth'))


def internal_has_permission(permission: str, db=None) -> bool:
    user = current_internal_user(db)
    if user:
        role = (user.role or 'quote_only').strip()
        perms = INTERNAL_ROLE_PERMISSIONS.get(role, {})
        if perms.get('all'):
            return True
        return bool(perms.get(permission))
    return bool(session.get('internal_auth'))


def require_internal_permission(permission):
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(*args, **kwargs):
            if not is_internal_logged_in():
                return redirect(url_for('internal_login', next=request.path))
            if not internal_has_permission(permission):
                flash('此帳號無此權限。')
                return redirect(url_for('internal_home') if internal_has_permission('quote_only') else url_for('internal_login'))
            return view_func(*args, **kwargs)
        return wrapper
    return decorator


def current_dealer(db=None):
    dealer_id = session.get('dealer_id')
    if not dealer_id:
        return None
    close_db = False
    if db is None:
        db = db_session()
        close_db = True
    try:
        return db.query(models.Dealer).filter(models.Dealer.id == dealer_id).first()
    finally:
        if close_db:
            db.close()




def dealer_permission_enabled(dealer, field_name: str) -> bool:
    if dealer is None:
        return False
    return bool(getattr(dealer, field_name, 1))


def require_internal_or_dealer_quote(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if is_internal_logged_in():
            return view_func(*args, **kwargs)
        db = db_session()
        try:
            dealer = current_dealer(db)
            if dealer and dealer_permission_enabled(dealer, 'can_create_quote'):
                return view_func(*args, **kwargs)
        finally:
            db.close()
        return redirect(url_for('dealer_login', next=request.path))
    return wrapper

def require_internal_login(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not is_internal_logged_in():
            return redirect(url_for('internal_login', next=request.path))
        return view_func(*args, **kwargs)
    return wrapper


def require_dealer_login(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get('dealer_id'):
            return redirect(url_for('dealer_login', next=request.path))
        return view_func(*args, **kwargs)
    return wrapper


@flask_app.context_processor
def inject_auth_state():
    db = None
    dealer_name = ''
    dealer_level = ''
    if session.get('dealer_id'):
        db = db_session()
        dealer = current_dealer(db)
        if dealer:
            dealer_name = dealer.name
            dealer_level = dealer.level or ''
        db.close()
    internal_user = current_internal_user(db) if False else None
    return {
        'is_internal_logged_in': is_internal_logged_in(),
        'is_dealer_logged_in': bool(session.get('dealer_id')),
        'dealer_name': dealer_name,
        'dealer_level': dealer_level,
        'internal_username': session.get('internal_username',''),
        'internal_display_name': session.get('internal_display_name',''),
        'internal_role': session.get('internal_role',''),
        'can_manage_all': bool(session.get('internal_role') == 'supervisor' or session.get('internal_auth')),
        'can_quote_only': bool(is_internal_logged_in()),
    }



def db_session():
    return SessionLocal()


def normalize_password_input(value: str) -> str:
    return (value or '').strip()[:10]


def normalize_tax_id_input(value: str) -> str:
    return re.sub(r'\D+', '', (value or '').strip())[:10]


def get_system_setting(db, key: str, default: str = '') -> str:
    row = db.query(models.SystemSetting).filter(models.SystemSetting.setting_key == key).first()
    return (row.setting_value if row and row.setting_value is not None else default)


def set_system_setting(db, key: str, value: str):
    row = db.query(models.SystemSetting).filter(models.SystemSetting.setting_key == key).first()
    normalized = value if value is not None else ''
    if row:
        row.setting_value = normalized
        row.updated_at = datetime.utcnow()
    else:
        db.add(models.SystemSetting(setting_key=key, setting_value=normalized, updated_at=datetime.utcnow()))


def get_internal_password(db=None) -> str:
    close_db = False
    if db is None:
        db = db_session()
        close_db = True
    try:
        value = get_system_setting(db, 'internal_password', DEFAULT_INTERNAL_PASSWORD)
        return normalize_password_input(value) or DEFAULT_INTERNAL_PASSWORD
    finally:
        if close_db:
            db.close()


def ensure_default_internal_users(db):
    legacy = db.query(models.InternalUser).filter(models.InternalUser.username == 'hqadmin').first()
    supervisor = db.query(models.InternalUser).filter(models.InternalUser.username == DEFAULT_SUPERVISOR_USERNAME).first()
    if legacy and not supervisor:
        legacy.username = DEFAULT_SUPERVISOR_USERNAME
        legacy.display_name = legacy.display_name or DEFAULT_SUPERVISOR_DISPLAY_NAME
        legacy.role = 'supervisor'
        legacy.is_active = 1
        legacy.updated_at = datetime.utcnow()
        db.commit()
        supervisor = legacy
    if not supervisor:
        db.add(models.InternalUser(username=DEFAULT_SUPERVISOR_USERNAME, password=DEFAULT_INTERNAL_PASSWORD, display_name=DEFAULT_SUPERVISOR_DISPLAY_NAME, role='supervisor', is_active=1, updated_at=datetime.utcnow()))
        db.commit()


def get_current_internal_user_or_abort(db):
    user = current_internal_user(db)
    if not user:
        flash('找不到目前登入帳號，請重新登入。')
        raise RuntimeError('missing current internal user')
    return user


def create_password_reset_request(db, account_type: str, account_identifier: str, display_name: str = ''):
    account_type = (account_type or '').strip()
    account_identifier = (account_identifier or '').strip()
    if not account_type or not account_identifier:
        return False, '請輸入帳號資料。'
    existing = db.query(models.PasswordResetRequest).filter(
        models.PasswordResetRequest.account_type == account_type,
        models.PasswordResetRequest.account_identifier == account_identifier,
        models.PasswordResetRequest.status == 'pending'
    ).first()
    if existing:
        return False, '此帳號已有待處理的忘記密碼申請。'
    db.add(models.PasswordResetRequest(
        account_type=account_type,
        account_identifier=account_identifier,
        display_name=(display_name or '').strip(),
        status='pending'
    ))
    db.commit()
    return True, '已送出忘記密碼申請，請由主管協助重設。'




with db_session() as _init_db:
    ensure_default_internal_users(_init_db)


def safe_filename(text_value: str) -> str:
    value = (text_value or "").strip()
    value = re.sub(r'[\\/:*?"<>|]+', '_', value)
    return value or "未填"


def quote_output_name(quote) -> str:
    return f"{quote.quote_date.strftime('%Y%m%d')} - {safe_filename(quote.customer_name)} - {safe_filename(quote.contact_name)} -報價單"


def get_image_src(product):
    image_path = (getattr(product, 'image_path', None) or '').strip()
    image_url = (getattr(product, 'image_url', None) or '').strip()
    model = safe_filename((getattr(product, 'model', None) or '').strip())

    if image_path:
        if image_path.startswith(('http://', 'https://', '/uploads/')):
            return image_path
        normalized = image_path.replace('\\', '/').lstrip('/')
        candidate = UPLOAD_DIR / normalized
        if candidate.exists():
            return f'/uploads/{normalized}'

    if image_url:
        return image_url

    if model:
        fallback_dirs = ['product_images', 'images']
        fallback_exts = ['.png', '.jpg', '.jpeg', '.webp', '.gif']
        for folder in fallback_dirs:
            for ext in fallback_exts:
                rel = f'{folder}/{model}{ext}'
                if (UPLOAD_DIR / rel).exists():
                    return f'/uploads/{rel}'
    return ""


def get_company_logo_src():
    for name in ['company_logo.png', 'company_logo.jpg', 'company_logo.jpeg']:
        p = UPLOAD_DIR / name
        if p.exists():
            return f'/uploads/{name}'
    return ''


def resolve_product_image_file(product):
    image_path = (getattr(product, 'image_path', None) or '').strip()
    if image_path.startswith('/uploads/'):
        rel = image_path[len('/uploads/'):].lstrip('/').replace('\\', '/')
        candidate = UPLOAD_DIR / rel
        if candidate.exists():
            return candidate, rel
    elif image_path:
        rel = image_path.lstrip('/').replace('\\', '/')
        candidate = UPLOAD_DIR / rel
        if candidate.exists():
            return candidate, rel

    model = safe_filename((getattr(product, 'model', None) or '').strip())
    if model:
        fallback_dirs = ['product_images', 'images']
        fallback_exts = ['.png', '.jpg', '.jpeg', '.webp', '.gif']
        for folder in fallback_dirs:
            for ext in fallback_exts:
                rel = f'{folder}/{model}{ext}'
                candidate = UPLOAD_DIR / rel
                if candidate.exists():
                    return candidate, rel
    return None, None


def ensure_original_backup(product, src_file: Path):
    if not src_file or not src_file.exists():
        return None
    ext = src_file.suffix.lower() or '.png'
    backup = ORIGINAL_IMAGE_DIR / f'product_{product.id}_original{ext}'
    if not backup.exists():
        shutil.copy2(src_file, backup)
    return backup


def normalize_image_target(product, preferred_suffix='.png'):
    src_file, rel = resolve_product_image_file(product)
    if rel:
        return UPLOAD_DIR / rel, rel
    model = safe_filename((getattr(product, 'model', None) or '').strip()) or f'product_{product.id}'
    rel = f'product_images/{model}{preferred_suffix}'
    return UPLOAD_DIR / rel, rel


def save_square_image(file_or_stream, target_path: Path, size=500):
    target_path.parent.mkdir(parents=True, exist_ok=True)
    image = Image.open(file_or_stream)
    image = ImageOps.exif_transpose(image)
    if image.mode not in ('RGB', 'RGBA'):
        image = image.convert('RGBA')
    square = Image.new('RGBA', (size, size), (255, 255, 255, 0))
    fit = ImageOps.contain(image, (size, size))
    x = (size - fit.width) // 2
    y = (size - fit.height) // 2
    square.paste(fit, (x, y), fit if fit.mode == 'RGBA' else None)
    if target_path.suffix.lower() in ['.jpg', '.jpeg']:
        square = square.convert('RGB')
        square.save(target_path, quality=95)
    else:
        square.save(target_path, format='PNG')
    return target_path


def get_original_backup_path(product):
    for ext in ['.png', '.jpg', '.jpeg', '.webp', '.gif']:
        p = ORIGINAL_IMAGE_DIR / f'product_{product.id}_original{ext}'
        if p.exists():
            return p
    return None

def get_cost_usd(product, usd_rate=32.5, rmb_rate=4.5):
    source_cost = safe_number(getattr(product, 'source_cost', 0) or 0)
    currency = (getattr(product, 'source_currency', 'TWD') or 'TWD').upper()
    if currency == 'USD':
        return round(source_cost, 2)
    if currency == 'RMB':
        return round(source_cost * (rmb_rate / usd_rate), 2) if usd_rate else 0
    return round(source_cost / usd_rate, 2) if usd_rate else 0


def get_cost_twd(product, usd_rate=32.5, rmb_rate=4.5):
    source_cost = safe_number(getattr(product, 'source_cost', 0) or 0)
    currency = (getattr(product, 'source_currency', 'TWD') or 'TWD').upper()
    if currency == 'USD':
        return round(source_cost * usd_rate, 0)
    if currency == 'RMB':
        return round(source_cost * rmb_rate, 0)
    return round(source_cost, 0)


def recompute_product_prices(product, rates, discount_map=None):
    usd_rate = safe_number(rates.get('USD', 32.5) or 32.5)
    rmb_rate = safe_number(rates.get('RMB', 4.5) or 4.5)
    product.shipping_usd = safe_number(getattr(product, 'shipping_usd', 0) or 0)
    product.duty_rate_pct = safe_number(getattr(product, 'duty_rate_pct', 0) or 0)
    product.outsourced_parts_fee_twd = round(safe_number(getattr(product, 'outsourced_parts_fee_twd', 0) or 0), 0)
    product.duty_cost_usd = calculate_duty_cost_usd(product.source_currency, product.source_cost, product.shipping_usd, product.duty_rate_pct, usd_rate, rmb_rate)
    product.duty_cost_twd = round(product.duty_cost_usd * usd_rate, 0) if product.duty_rate_pct > 0 else round(safe_number(getattr(product, 'duty_cost_twd', 0)), 0)
    product.final_cost_twd = calculate_final_cost_twd(product.source_currency, product.source_cost, usd_rate, rmb_rate, product.duty_cost_twd, product.shipping_usd, product.duty_rate_pct, product.outsourced_parts_fee_twd)
    specials = {
        'special_market_price_ratio': getattr(product, 'special_market_price_ratio', 1) or 1,
        'special_market_min_ratio': getattr(product, 'special_market_min_ratio', 1) or 1,
        'special_designer_ratio': getattr(product, 'special_designer_ratio', 1) or 1,
        'special_dealer_lv1_ratio': getattr(product, 'special_dealer_lv1_ratio', 1) or 1,
        'special_dealer_lv2_ratio': getattr(product, 'special_dealer_lv2_ratio', 1) or 1,
        'special_branch_ratio': getattr(product, 'special_branch_ratio', 1) or 1,
        'special_master_ratio': getattr(product, 'special_master_ratio', 1) or 1,
    }
    effective_discount_map = dict(discount_map or DEFAULT_DISCOUNT_RATIOS)
    effective_discount_map.update(get_product_discount_ratio_map(product, effective_discount_map))
    built = build_price_levels(product.final_cost_twd, specials, effective_discount_map)
    product.market_price = round(safe_number(built['market_price']), 0)
    product.market_min_price = round(safe_number(built['market_min_price']), 0)
    product.designer_price = round(safe_number(built['designer_price']), 0)
    product.dealer_lv1_price = round(safe_number(built['dealer_lv1_price']), 0)
    product.dealer_lv2_price = round(safe_number(built['dealer_lv2_price']), 0)
    product.branch_price = round(safe_number(built['branch_price']), 0)
    product.master_dealer_price = round(safe_number(built['master_dealer_price']), 0)
    return product


def ensure_sqlite_columns():
    statements = [
        "ALTER TABLE products ADD COLUMN image_path TEXT",
        "ALTER TABLE products ADD COLUMN planning_fee_twd FLOAT DEFAULT 0",
        "ALTER TABLE products ADD COLUMN setup_fee_twd FLOAT DEFAULT 0",
        "ALTER TABLE products ADD COLUMN shipping_usd FLOAT DEFAULT 0",
        "ALTER TABLE products ADD COLUMN duty_rate_pct FLOAT DEFAULT 0",
        "ALTER TABLE products ADD COLUMN duty_cost_usd FLOAT DEFAULT 0",
        "ALTER TABLE products ADD COLUMN outsourced_parts_fee_twd FLOAT DEFAULT 0",
        "ALTER TABLE products ADD COLUMN special_market_price_ratio FLOAT DEFAULT 1",
        "ALTER TABLE products ADD COLUMN special_market_min_ratio FLOAT DEFAULT 1",
        "ALTER TABLE products ADD COLUMN special_designer_ratio FLOAT DEFAULT 1",
        "ALTER TABLE products ADD COLUMN special_dealer_lv1_ratio FLOAT DEFAULT 1",
        "ALTER TABLE products ADD COLUMN special_dealer_lv2_ratio FLOAT DEFAULT 1",
        "ALTER TABLE products ADD COLUMN special_branch_ratio FLOAT DEFAULT 1",
        "ALTER TABLE products ADD COLUMN special_master_ratio FLOAT DEFAULT 1",
        "ALTER TABLE products ADD COLUMN market_min_discount_ratio FLOAT DEFAULT 95",
        "ALTER TABLE products ADD COLUMN designer_discount_ratio FLOAT DEFAULT 90",
        "ALTER TABLE products ADD COLUMN dealer_lv1_discount_ratio FLOAT DEFAULT 85",
        "ALTER TABLE products ADD COLUMN dealer_lv2_discount_ratio FLOAT DEFAULT 80",
        "ALTER TABLE products ADD COLUMN branch_discount_ratio FLOAT DEFAULT 75",
        "ALTER TABLE products ADD COLUMN master_discount_ratio FLOAT DEFAULT 70",
        "ALTER TABLE products ADD COLUMN status VARCHAR(50)",
        "ALTER TABLE dealers ADD COLUMN access_key VARCHAR(50)",
        "ALTER TABLE dealers ADD COLUMN can_view_products INTEGER DEFAULT 1",
        "ALTER TABLE dealers ADD COLUMN can_export_prices INTEGER DEFAULT 1",
        "ALTER TABLE dealers ADD COLUMN can_create_quote INTEGER DEFAULT 1",
        "ALTER TABLE dealers ADD COLUMN sales_owner VARCHAR(255)",
        "ALTER TABLE dealers ADD COLUMN payment_method VARCHAR(100)",
        "ALTER TABLE dealers ADD COLUMN note TEXT",
        "ALTER TABLE dealers ADD COLUMN signed_month VARCHAR(20)",
        "ALTER TABLE quotes ADD COLUMN attn VARCHAR(255)",
        "ALTER TABLE quotes ADD COLUMN sales_name VARCHAR(255)",
        "ALTER TABLE quotes ADD COLUMN sales_phone VARCHAR(100)",
        "ALTER TABLE quotes ADD COLUMN sales_email VARCHAR(255)",
        "ALTER TABLE quotes ADD COLUMN currency VARCHAR(20) DEFAULT 'NTD'",
        "ALTER TABLE quotes ADD COLUMN product_subtotal FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN planning_fee_total FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN setup_fee_total FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN dispatch_fee FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN planning_multiplier FLOAT DEFAULT 1",
        "ALTER TABLE quotes ADD COLUMN setup_multiplier FLOAT DEFAULT 1",
        "ALTER TABLE quotes ADD COLUMN dispatch_label VARCHAR(100)",
        "ALTER TABLE quotes ADD COLUMN lock_install_qty INTEGER DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN lock_install_unit_price FLOAT DEFAULT 3800",
        "ALTER TABLE quotes ADD COLUMN lock_install_fee FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN curtain_install_qty FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN curtain_install_unit VARCHAR(50) DEFAULT '式'",
        "ALTER TABLE quotes ADD COLUMN curtain_install_amount FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN curtain_type VARCHAR(50)",
        "ALTER TABLE quotes ADD COLUMN curtain_motor_price FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN curtain_track_unit_price FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN curtain_track_length FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN curtain_fabric_width FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN curtain_fabric_height FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN curtain_fabric_unit_price FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN curtain_note TEXT",
        "ALTER TABLE quotes ADD COLUMN curtain_rows_json TEXT",
        "ALTER TABLE quotes ADD COLUMN weak_current_qty FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN weak_current_unit VARCHAR(50) DEFAULT '式'",
        "ALTER TABLE quotes ADD COLUMN weak_current_amount FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN hardware_qty FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN hardware_unit VARCHAR(50) DEFAULT '式'",
        "ALTER TABLE quotes ADD COLUMN hardware_amount FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN water_elec_qty FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN water_elec_unit VARCHAR(50) DEFAULT '式'",
        "ALTER TABLE quotes ADD COLUMN water_elec_amount FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN custom_fee_json TEXT",
        "ALTER TABLE quotes ADD COLUMN gross_profit_rate FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN negotiated_total FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN negotiated_discount_pct FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN deposit_1 FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN deposit_2 FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN deposit_3 FLOAT DEFAULT 0",
        "ALTER TABLE quotes ADD COLUMN payment_scheme VARCHAR(50)",
        "ALTER TABLE quote_items ADD COLUMN cost_total_twd FLOAT DEFAULT 0",
        "ALTER TABLE quote_items ADD COLUMN planning_fee_twd FLOAT DEFAULT 0",
        "ALTER TABLE quote_items ADD COLUMN setup_fee_twd FLOAT DEFAULT 0",
        "ALTER TABLE quote_items ADD COLUMN image_path TEXT",
    ]
    with engine.begin() as conn:
        for sql in statements:
            try:
                conn.execute(text(sql))
            except Exception:
                pass


def startup_seed():
    ensure_sqlite_columns()
    db = db_session()
    try:
        ensure_default_rates(db)
        ensure_default_price_settings(db)
        ensure_default_sales_people(db)
        ensure_default_dealer_access_keys(db)
        ensure_default_dealer_permissions(db)
        ensure_default_system_settings(db)
    finally:
        db.close()



def ensure_default_system_settings(db):
    if not db.query(models.SystemSetting).filter(models.SystemSetting.setting_key == 'internal_password').first():
        db.add(models.SystemSetting(setting_key='internal_password', setting_value=DEFAULT_INTERNAL_PASSWORD, updated_at=datetime.utcnow()))
        db.commit()




def ensure_default_sales_people(db):
    defaults = [
        ("未設定業務", "", ""),
    ]
    for name, phone, email in defaults:
        row = db.query(models.SalesPerson).filter(models.SalesPerson.name == name).first()
        if not row:
            db.add(models.SalesPerson(name=name, phone=phone, email=email))
    db.commit()


def ensure_default_dealer_access_keys(db):
    def digits(v):
        return ''.join(ch for ch in (v or '') if ch.isdigit())
    changed = False
    for dealer in db.query(models.Dealer).all():
        if not getattr(dealer, 'access_key', None):
            key = digits(getattr(dealer, 'tax_id', ''))[-4:] or digits(getattr(dealer, 'phone', ''))[-4:] or '0000'
            dealer.access_key = key
            changed = True
    if changed:
        db.commit()


def ensure_default_dealer_permissions(db):
    changed = False
    for dealer in db.query(models.Dealer).all():
        for field in ['can_view_products','can_export_prices','can_create_quote']:
            if getattr(dealer, field, None) is None:
                setattr(dealer, field, 1)
                changed = True
    if changed:
        db.commit()




def ensure_default_options(db):
    default_categories = ['主機', '燈具類', '開關類', '門鎖類', '窗簾類', '感測器', '控制器', '馬達類']
    default_statuses = ['一般', '停產', '促銷', '特殊品項']
    existing_categories = {x.name for x in db.query(models.CategoryOption).all()}
    existing_statuses = {x.name for x in db.query(models.StatusOption).all()}
    changed = False
    for name in default_categories:
        if name not in existing_categories:
            db.add(models.CategoryOption(name=name))
            changed = True
    for name in default_statuses:
        if name not in existing_statuses:
            db.add(models.StatusOption(name=name))
            changed = True
    if changed:
        db.commit()


def get_option_lists(db):
    categories = [r.name for r in db.query(models.CategoryOption).order_by(models.CategoryOption.name).all()]
    statuses = [r.name for r in db.query(models.StatusOption).order_by(models.StatusOption.name).all()]
    # fallback from product data
    for value in [x[0] for x in db.query(models.Product.category).distinct().all() if x[0]]:
        if value not in categories:
            categories.append(value)
    for value in [x[0] for x in db.query(models.Product.status).distinct().all() if x[0]]:
        if value not in statuses:
            statuses.append(value)
    return sorted(categories), sorted(statuses)


def record_product_change(db, product_id, details, changed_by='internal'):
    if not details:
        return
    detail_text = '\n'.join(details)[:4000]
    db.add(models.ProductChangeLog(product_id=product_id, changed_by=changed_by, detail=detail_text))
    db.commit()


def persist_ui_state(key, default=''):
    return request.args.get(key, default)


startup_seed()
with db_session() as _db:
    ensure_default_options(_db)

def cleanup_old_logs(db):
    cutoff = datetime.utcnow().timestamp() - 730 * 24 * 60 * 60
    cutoff_dt = datetime.utcfromtimestamp(cutoff)
    db.query(models.AuditLog).filter(models.AuditLog.created_at < cutoff_dt).delete()
    db.commit()


def log_event(db, event_type: str, target_type: str, detail: str, target_id: str = ""):
    db.add(models.AuditLog(event_type=event_type, target_type=target_type, target_id=target_id, detail=detail[:2000]))
    db.commit()
    cleanup_old_logs(db)


PRODUCT_DISCOUNT_ATTR_MAP = {
    "市場最低價": "market_min_discount_ratio",
    "設計師價": "designer_discount_ratio",
    "一級經銷商": "dealer_lv1_discount_ratio",
    "二級經銷商": "dealer_lv2_discount_ratio",
    "分公司": "branch_discount_ratio",
    "總經銷商": "master_discount_ratio",
}

PRODUCT_PRICE_ATTR_MAP = {
    "市場最低價": "market_min_price",
    "設計師價": "designer_price",
    "一級經銷商": "dealer_lv1_price",
    "二級經銷商": "dealer_lv2_price",
    "分公司": "branch_price",
    "總經銷商": "master_dealer_price",
}

PRODUCT_SPECIAL_ATTR_MAP = {
    "市場報價": "special_market_price_ratio",
    "市場最低價": "special_market_min_ratio",
    "設計師價": "special_designer_ratio",
    "一級經銷商": "special_dealer_lv1_ratio",
    "二級經銷商": "special_dealer_lv2_ratio",
    "分公司": "special_branch_ratio",
    "總經銷商": "special_master_ratio",
}

def derive_product_discount_ratio(product, level_name, default_ratio=0):
    attr = PRODUCT_DISCOUNT_ATTR_MAP.get(level_name)
    if attr:
        raw = safe_number(getattr(product, attr, 0) or 0)
        if raw > 0:
            return normalize_percentage_value(raw)
    market = safe_number(getattr(product, 'market_price', 0) or 0)
    child_attr = PRODUCT_PRICE_ATTR_MAP.get(level_name)
    special_attr = PRODUCT_SPECIAL_ATTR_MAP.get(level_name)
    child = safe_number(getattr(product, child_attr, 0) or 0) if child_attr else 0
    special = safe_number(getattr(product, special_attr, 1) or 1) if special_attr else 1
    if market > 0 and child > 0 and special > 0:
        return round(child / market / special * 100.0, 4)
    return normalize_percentage_value(default_ratio)

def get_product_discount_ratio_map(product, default_map=None):
    default_map = default_map or DEFAULT_DISCOUNT_RATIOS
    result = {}
    for level_name in PRODUCT_DISCOUNT_ATTR_MAP:
        result[level_name] = derive_product_discount_ratio(product, level_name, default_map.get(level_name, 0))
    return result

def average_product_level_ratios(db):
    products = db.query(models.Product).all()
    default_map = get_price_setting_map(db)
    result = {}
    for level_name in PRODUCT_DISCOUNT_ATTR_MAP:
        values = []
        for product in products:
            ratio = derive_product_discount_ratio(product, level_name, default_map.get(level_name, 0))
            if ratio > 0:
                values.append(ratio)
        result[level_name] = round(sum(values) / len(values), 2) if values else 0.0
    return result


def get_product_discount_display_map(db):
    discount_map = get_price_setting_map(db)
    return {k: normalize_percentage_value(v) for k, v in discount_map.items() if k != '市場報價'}


def selected_export_labels(selected_levels):
    mapping = {
        'market_price': '市場報價',
        'market_min_price': '市場最低價',
        'designer_price': '設計師價',
        'dealer_lv1_price': '一級經銷商',
        'dealer_lv2_price': '二級經銷商',
        'branch_price': '分公司',
        'master_dealer_price': '總經銷商',
    }
    return [mapping[x] for x in selected_levels if x in mapping]





def build_product_export_workbook(products, selected_levels, include_fields=None):
    include_fields = include_fields or {}
    if not selected_levels:
        selected_levels = ['market_price']

    wb = Workbook()
    ws = wb.active
    ws.title = '商品總表'
    sys_ws = wb.create_sheet('系統匯入資料')

    ws['A1'] = '公式說明：主表顯示結果，系統匯入資料保留完整公式。狀態=停產優先，否則特殊倍數>=1.05為特殊品項、<=0.95為促銷，其餘為一般。'
    sys_ws['A1'] = '公式說明：M=ROUND(IF(L<1,(J+K)*L+(J+K),(J+K)*(L/100)+(J+K)),2)；O=ROUND(M*I+N,0)。'

    # after removing P欄 成本(USD)
    level_meta = {
        'market_price': ('市場報價', 'Y', None),
        'market_min_price': ('市場最低價', 'Z', 'AF'),
        'designer_price': ('設計師價', 'AA', 'AG'),
        'dealer_lv1_price': ('一級經銷商', 'AB', 'AH'),
        'dealer_lv2_price': ('二級經銷商', 'AC', 'AI'),
        'branch_price': ('分公司', 'AD', 'AJ'),
        'master_dealer_price': ('總經銷商', 'AE', 'AK'),
    }

    headers = []
    if include_fields.get('category', True): headers.append('類別')
    headers.append('狀態')
    headers.extend(['型號','品名','單位'])
    if include_fields.get('description', True): headers.append('產品描述')
    if include_fields.get('image', True): headers.append('圖片')
    if include_fields.get('cost_twd', True): headers.append('成本(TWD)')
    if include_fields.get('special_ratio', False): headers.append('特殊倍數')
    if include_fields.get('outsourced_fee', False): headers.append('其他加購（TWD）')
    if include_fields.get('duty_rate_pct', False): headers.append('關稅 / 其他成本加成（%）')
    if include_fields.get('duty_cost_usd', False): headers.append('加乘後金額（USD）')
    if include_fields.get('planning_fee', False): headers.append('規劃費')
    if include_fields.get('setup_fee', False): headers.append('設定費')

    for lvl in selected_levels:
        label, _, disc_col = level_meta[lvl]
        headers.append(label)
        if disc_col:
            headers.append(f'{label}折扣比例(相對市場報價%)')
        if include_fields.get('profit_rate', False):
            headers.append(f'{label}利潤率')
    headers.append('備註')
    ws.append([])
    ws.append([])
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)

    sys_headers = [
        '類別','狀態','型號','品名','單位','產品描述','圖片','來源幣別','匯率','來源成本(USD)','運費(USD)',
        '關稅/其他成本加成(%)','加乘後金額(USD)','其他加購(TWD)','成本(TWD)',
        '規劃費','設定費',
        '特殊倍數_市場價','特殊倍數_市場最低價','特殊倍數_設計師價','特殊倍數_一級經銷商','特殊倍數_二級經銷商','特殊倍數_分公司','特殊倍數_總經銷商',
        '市場報價','市場最低價','設計師價','一級經銷商','二級經銷商','分公司','總經銷商',
        '市場最低價折扣比例','設計師價折扣比例','一級經銷商折扣比例','二級經銷商折扣比例','分公司折扣比例','總經銷商折扣比例',
        '備註'
    ]
    sys_ws.append([])
    sys_ws.append([])
    for c, h in enumerate(sys_headers, start=1):
        sys_ws.cell(row=3, column=c, value=h)

    image_col_main = headers.index('圖片') + 1 if '圖片' in headers else None

    def status_by_system(p):
        raw = (getattr(p, 'status', '') or '').strip()
        if '停產' in raw:
            return '停產'
        ratios = [
            round3(safe_number(getattr(p, 'special_market_price_ratio', 1) or 1), 1),
            round3(safe_number(getattr(p, 'special_market_min_ratio', 1) or 1), 1),
            round3(safe_number(getattr(p, 'special_designer_ratio', 1) or 1), 1),
            round3(safe_number(getattr(p, 'special_dealer_lv1_ratio', 1) or 1), 1),
            round3(safe_number(getattr(p, 'special_dealer_lv2_ratio', 1) or 1), 1),
            round3(safe_number(getattr(p, 'special_branch_ratio', 1) or 1), 1),
            round3(safe_number(getattr(p, 'special_master_ratio', 1) or 1), 1),
        ]
        if any(r >= 1.05 for r in ratios):
            return '特殊品項'
        if any(r <= 0.95 for r in ratios):
            return '促銷'
        return '一般'

    def rel_ratio(price, market_price):
        return round((price / market_price), 4) if market_price else 0

    main_row = 4
    sys_row = 4

    for p in products:
        usd_rate = safe_number(getattr(p, 'usd_rate', 32.5) or 32.5)
        src_cost = safe_number(getattr(p, 'source_cost', 0))
        ship_usd = safe_number(getattr(p, 'shipping_usd', 0))
        duty_pct = safe_number(getattr(p, 'duty_rate_pct', 0))
        outsourced = safe_number(getattr(p, 'outsourced_parts_fee_twd', 0))
        planning = round(safe_number(getattr(p, 'planning_fee_twd', 0)), 0)
        setup = round(safe_number(getattr(p, 'setup_fee_twd', 0)), 0)
        market = round(safe_number(getattr(p, 'market_price', 0)), 0)
        minp = round(safe_number(getattr(p, 'market_min_price', 0)), 0)
        designer = round(safe_number(getattr(p, 'designer_price', 0)), 0)
        lv1 = round(safe_number(getattr(p, 'dealer_lv1_price', 0)), 0)
        lv2 = round(safe_number(getattr(p, 'dealer_lv2_price', 0)), 0)
        branch = round(safe_number(getattr(p, 'branch_price', 0)), 0)
        master = round(safe_number(getattr(p, 'master_dealer_price', 0)), 0)

        sys_vals = [
            p.category or '', status_by_system(p), p.model, p.name, p.unit or '台', p.description or '', get_image_src(p) or '',
            p.source_currency or 'USD', usd_rate, src_cost, ship_usd, duty_pct, None, outsourced, None,
            planning, setup,
            round3(safe_number(getattr(p, 'special_market_price_ratio', 1) or 1), 1),
            round3(safe_number(getattr(p, 'special_market_min_ratio', 1) or 1), 1),
            round3(safe_number(getattr(p, 'special_designer_ratio', 1) or 1), 1),
            round3(safe_number(getattr(p, 'special_dealer_lv1_ratio', 1) or 1), 1),
            round3(safe_number(getattr(p, 'special_dealer_lv2_ratio', 1) or 1), 1),
            round3(safe_number(getattr(p, 'special_branch_ratio', 1) or 1), 1),
            round3(safe_number(getattr(p, 'special_master_ratio', 1) or 1), 1),
            None, None, None, None, None, None, None,
            rel_ratio(minp, market), rel_ratio(designer, market), rel_ratio(lv1, market), rel_ratio(lv2, market), rel_ratio(branch, market), rel_ratio(master, market),
            p.note or ''
        ]
        for c, v in enumerate(sys_vals, start=1):
            sys_ws.cell(row=sys_row, column=c, value=v)

        sys_ws[f'M{sys_row}'] = f'=ROUND(IF(L{sys_row}<1,(J{sys_row}+K{sys_row})*L{sys_row}+(J{sys_row}+K{sys_row}),(J{sys_row}+K{sys_row})*(L{sys_row}/100)+(J{sys_row}+K{sys_row})),2)'
        sys_ws[f'O{sys_row}'] = f'=ROUND(M{sys_row}*I{sys_row}+N{sys_row},0)'
        sys_ws[f'Y{sys_row}'] = f'=ROUND(O{sys_row}*R{sys_row},0)'
        sys_ws[f'Z{sys_row}'] = f'=ROUND(O{sys_row}*S{sys_row}*AF{sys_row},0)'
        sys_ws[f'AA{sys_row}'] = f'=ROUND(O{sys_row}*T{sys_row}*AG{sys_row},0)'
        sys_ws[f'AB{sys_row}'] = f'=ROUND(O{sys_row}*U{sys_row}*AH{sys_row},0)'
        sys_ws[f'AC{sys_row}'] = f'=ROUND(O{sys_row}*V{sys_row}*AI{sys_row},0)'
        sys_ws[f'AD{sys_row}'] = f'=ROUND(O{sys_row}*W{sys_row}*AJ{sys_row},0)'
        sys_ws[f'AE{sys_row}'] = f'=ROUND(O{sys_row}*X{sys_row}*AK{sys_row},0)'

        row = []
        if include_fields.get('category', True): row.append(f"='系統匯入資料'!A{sys_row}")
        row.append(f"='系統匯入資料'!B{sys_row}")
        row.extend([f"='系統匯入資料'!C{sys_row}", f"='系統匯入資料'!D{sys_row}", f"='系統匯入資料'!E{sys_row}"])
        if include_fields.get('description', True): row.append(f"='系統匯入資料'!F{sys_row}")
        if include_fields.get('image', True): row.append('')
        if include_fields.get('cost_twd', True): row.append(f"='系統匯入資料'!O{sys_row}")
        if include_fields.get('special_ratio', False):
            row.append(f'="市場 "&TEXT(R{sys_row},"0.000")&"/最低 "&TEXT(S{sys_row},"0.000")&"/設計 "&TEXT(T{sys_row},"0.000")&"/一級 "&TEXT(U{sys_row},"0.000")&"/二級 "&TEXT(V{sys_row},"0.000")&"/分公司 "&TEXT(W{sys_row},"0.000")&"/總經 "&TEXT(X{sys_row},"0.000")')
        if include_fields.get('outsourced_fee', False): row.append(f"='系統匯入資料'!N{sys_row}")
        if include_fields.get('duty_rate_pct', False): row.append(f"='系統匯入資料'!L{sys_row}")
        if include_fields.get('duty_cost_usd', False): row.append(f"='系統匯入資料'!M{sys_row}")
        if include_fields.get('planning_fee', False): row.append(f"='系統匯入資料'!P{sys_row}")
        if include_fields.get('setup_fee', False): row.append(f"='系統匯入資料'!Q{sys_row}")

        profit_positions = []
        for lvl in selected_levels:
            label, result_col, disc_col = level_meta[lvl]
            row.append(f"='系統匯入資料'!{result_col}{sys_row}")
            if disc_col:
                row.append(f"='系統匯入資料'!{disc_col}{sys_row}")
            if include_fields.get('profit_rate', False):
                row.append(None)
                profit_positions.append(len(row))
        row.append(f"='系統匯入資料'!AM{sys_row}")

        for c, val in enumerate(row, start=1):
            ws.cell(row=main_row, column=c, value=val)

        if include_fields.get('profit_rate', False) and '成本(TWD)' in headers:
            cost_col_idx = headers.index('成本(TWD)') + 1
            for pos in profit_positions:
                price_col_idx = pos - 1
                ws.cell(row=main_row, column=pos, value=f'=IFERROR(ROUND((1-({get_column_letter(cost_col_idx)}{main_row}/{get_column_letter(price_col_idx)}{main_row}))*100,1)&"%","0.0%")')

        if image_col_main:
            ref = get_image_src(p)
            local = None
            if ref and ref.startswith('/uploads/'):
                cand = BASE_DIR / ref.lstrip('/')
                if cand.exists():
                    local = cand
            elif ref:
                cand = Path(ref)
                if cand.exists():
                    local = cand
            if local:
                try:
                    img = XLImage(str(local))
                    img.width = 48
                    img.height = 48
                    ws.add_image(img, ws.cell(row=main_row, column=image_col_main).coordinate)
                    ws.row_dimensions[main_row].height = 42
                except Exception:
                    pass

        main_row += 1
        sys_row += 1

    for sheet in [ws, sys_ws]:
        for col_cells in sheet.columns:
            letter = get_column_letter(col_cells[0].column)
            sample = list(col_cells)[:120]
            max_len = max(len(str(c.value or '')) for c in sample) if sample else 12
            sheet.column_dimensions[letter].width = min(max(max_len + 3, 12), 38)
    if image_col_main:
        ws.column_dimensions[get_column_letter(image_col_main)].width = 14

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def serialize_product(product, discount_map, usd_rate=32.5, rmb_rate=4.5):
    cost_twd = get_cost_twd(product, usd_rate, rmb_rate)
    cost_usd = get_cost_usd(product, usd_rate, rmb_rate)
    return {
        "id": product.id,
        "category": product.category or "",
        "model": product.model,
        "name": product.name,
        "unit": product.unit or "台",
        "description": product.description or "",
        "note": product.note or "",
        "status": product.status or "",
        "cost": round(cost_twd, 0),
        "cost_usd": round(cost_usd, 2),
        "shipping_usd": round(safe_number(getattr(product, "shipping_usd", 0)), 2),
        "duty_rate_pct": round(safe_number(getattr(product, "duty_rate_pct", 0)), 2),
        "duty_cost_usd": round(safe_number(getattr(product, "duty_cost_usd", 0)), 2),
        "planning_fee": round(safe_number(product.planning_fee_twd), 0),
        "setup_fee": round(safe_number(product.setup_fee_twd), 0),
        "prices": {level: round(get_price_by_level(product, level, discount_map), 0) for level in PRICE_LEVELS},
        "image": get_image_src(product),
        "image_url": get_image_src(product),
    }


def dispatch_map():
    return {label: amount for label, amount in DISPATCH_OPTIONS}


def build_auto_note(selected_products):
    product_notes = []
    for product in selected_products:
        note = (product.note or "").strip()
        if note and note not in product_notes:
            product_notes.append(note)
    if product_notes:
        return "商品備註：\n" + "\n".join(f"- {n}" for n in product_notes) + "\n\n" + DEFAULT_NOTE
    return DEFAULT_NOTE




def build_display_items(quote):
    items = list(quote.items)
    fee_rows = []
    try:
        curtain_rows = json.loads(getattr(quote, 'curtain_rows_json', '') or '[]')
    except Exception:
        curtain_rows = []
    for row in curtain_rows:
        qty = safe_int(row.get('qty') or 0)
        unit_price = round(safe_number(row.get('unit_price') or row.get('line_total') or 0), 0)
        line_total = round(safe_number(row.get('line_total') or 0), 0)
        if qty <= 0 or unit_price <= 0:
            continue
        space = (row.get('space') or '').strip()
        ctype = (row.get('type') or '窗簾').strip()
        track_length = safe_int(row.get('track_length') or 0)
        height = safe_int(row.get('height') or 0)
        product_name = f"{space}-{ctype} {track_length}cm*{height}cm" if space else f"{ctype} {track_length}cm*{height}cm"
        fee_rows.append({
            "image_path": "", "model": "", "product_name": product_name,
            "qty": qty, "unit": "道",
            "unit_price_twd": unit_price,
            "line_total_twd": line_total if line_total else round(unit_price * qty, 0),
            "note": row.get('note') or ''
        })
    if round(getattr(quote, 'planning_fee_total', 0) or 0):
        fee_rows.append({"image_path": "", "model": "", "product_name": "規劃費", "qty": 1, "unit": "式", "unit_price_twd": round(quote.planning_fee_total or 0), "line_total_twd": round(quote.planning_fee_total or 0), "note": ''})
    if round(getattr(quote, 'setup_fee_total', 0) or 0):
        fee_rows.append({"image_path": "", "model": "", "product_name": "設定費", "qty": 1, "unit": "式", "unit_price_twd": round(quote.setup_fee_total or 0), "line_total_twd": round(quote.setup_fee_total or 0), "note": ''})
    if round(getattr(quote, 'dispatch_fee', 0) or 0):
        fee_rows.append({"image_path": "", "model": "", "product_name": "派工費", "qty": 1, "unit": "式", "unit_price_twd": round(quote.dispatch_fee or 0), "line_total_twd": round(quote.dispatch_fee or 0), "note": ""})
    if round(getattr(quote, 'lock_install_fee', 0) or 0):
        fee_rows.append({"image_path": "", "model": "", "product_name": "門鎖施工費", "qty": getattr(quote,'lock_install_qty',0) or 0, "unit": "把", "unit_price_twd": round(getattr(quote,'lock_install_unit_price',0) or 0), "line_total_twd": round(getattr(quote,'lock_install_fee',0) or 0), "note": ""})
    for key, label in [('curtain_install','電動窗簾安裝費'),('weak_current','弱電費用'),('hardware','五金材料'),('water_elec','水電費用')]:
        amount = round(getattr(quote, f'{key}_amount', 0) or 0)
        if amount:
            fee_rows.append({"image_path": "", "model": "", "product_name": label, "qty": getattr(quote,f'{key}_qty',1) or 1, "unit": getattr(quote,f'{key}_unit','式') or '式', "unit_price_twd": amount if (getattr(quote,f'{key}_qty',1) or 1)==1 else round(amount / (getattr(quote,f'{key}_qty',1) or 1),0), "line_total_twd": amount, "note": ""})
    try:
        custom_rows = json.loads(getattr(quote, 'custom_fee_json', '') or '[]')
    except Exception:
        custom_rows = []
    for row in custom_rows:
        total = round(safe_number(row.get('total') or 0), 0)
        if total:
            fee_rows.append({
                "image_path": "", "model": "", "product_name": row.get('name') or '自訂工費',
                "qty": row.get('qty') or 0, "unit": row.get('unit') or '式',
                "unit_price_twd": round(safe_number(row.get('unit_price') or 0), 0),
                "line_total_twd": total, "note": ""
            })
    return items + fee_rows

def parse_custom_fee_rows(form):
    names = form.getlist("custom_fee_name") or form.getlist("custom_fee_name[]")
    qtys = form.getlist("custom_fee_qty") or form.getlist("custom_fee_qty[]")
    units = form.getlist("custom_fee_unit") or form.getlist("custom_fee_unit[]")
    unit_prices = form.getlist("custom_fee_unit_price") or form.getlist("custom_fee_unit_price[]")
    totals = form.getlist("custom_fee_total") or form.getlist("custom_fee_total[]")
    notes = form.getlist("custom_fee_note") or form.getlist("custom_fee_note[]")
    rows = []
    for i, name in enumerate(names):
        name = (name or "").strip()
        qty = safe_int(qtys[i] if i < len(qtys) else 0)
        unit = (units[i] if i < len(units) else '') or '式'
        unit_price = round(safe_number(unit_prices[i] if i < len(unit_prices) else 0), 0)
        total = round(safe_number(totals[i] if i < len(totals) else 0), 0)
        note = (notes[i] if i < len(notes) else '').strip()
        if not name and not qty and not unit_price and not total and not note:
            continue
        if not total and qty and unit_price:
            total = round(qty * unit_price, 0)
        rows.append({
            "name": name or '自訂工費',
            "qty": qty or 0,
            "unit": unit,
            "unit_price": unit_price,
            "total": total,
            "note": note,
        })
    return rows


def parse_curtain_rows(form):
    spaces = form.getlist("curtain_space[]")
    types = form.getlist("curtain_type[]")
    qtys = form.getlist("curtain_tracks[]")
    motor_prices = form.getlist("curtain_motor_price[]")
    track_prices = form.getlist("curtain_track_unit_price[]")
    track_lengths = form.getlist("curtain_track_length[]")
    widths = form.getlist("curtain_width[]")
    heights = form.getlist("curtain_height[]")
    cloth_prices = form.getlist("curtain_cloth_price[]")
    totals = form.getlist("curtain_total[]")
    notes = form.getlist("curtain_note[]")
    rows = []
    max_len = max([len(spaces), len(types), len(qtys), len(motor_prices), len(track_prices), len(track_lengths), len(widths), len(heights), len(cloth_prices), len(totals), len(notes)] or [0])
    for i in range(max_len):
        space = (spaces[i] if i < len(spaces) else '').strip()
        ctype = (types[i] if i < len(types) else '').strip()
        qty = max(safe_int(qtys[i] if i < len(qtys) else 0), 0)
        motor_price = max(safe_number(motor_prices[i] if i < len(motor_prices) else 0),0)
        track_price = max(safe_number(track_prices[i] if i < len(track_prices) else 0),0)
        track_length = max(safe_number(track_lengths[i] if i < len(track_lengths) else 0),0)
        width = max(safe_number(widths[i] if i < len(widths) else 0),0)
        height = max(safe_number(heights[i] if i < len(heights) else 0),0)
        cloth_price = max(safe_number(cloth_prices[i] if i < len(cloth_prices) else 0),0)
        total = round(max(safe_number(totals[i] if i < len(totals) else 0),0), 0)
        note = (notes[i] if i < len(notes) else '').strip()
        if not any([space, ctype, qty, motor_price, track_price, track_length, width, height, cloth_price, total, note]):
            continue
        track_cost = math.ceil(track_length / 30.0) * track_price if track_length > 0 and track_price > 0 else 0
        cloth_cost = math.ceil((width * height) / 900.0) * cloth_price if width > 0 and height > 0 and cloth_price > 0 else 0
        computed_total = round(motor_price + track_cost + cloth_cost, 0)
        if total != computed_total:
            total = computed_total
        rows.append({
            "space": space,
            "type": ctype or '窗簾',
            "qty": qty,
            "motor_price": motor_price,
            "track_unit_price": track_price,
            "track_length": track_length,
            "width": width,
            "height": height,
            "cloth_price": cloth_price,
            "unit_price": total,
            "line_total": round(total * qty, 0),
            "note": note,
        })
    return rows


def quote_summary_from_request(db, form, product_ids, qtys):
    dealer_id = form.get('dealer_id')
    dealer = db.query(models.Dealer).filter(models.Dealer.id == int(dealer_id)).first() if dealer_id else None
    selected_level = form.get('price_level') or (dealer.level if dealer and dealer.level else '市場報價')
    discount_map = get_price_setting_map(db)
    product_map = {p.id: p for p in db.query(models.Product).all()}

    planning_multiplier = safe_number(form.get('planning_multiplier') or 1) or 1
    setup_multiplier = safe_number(form.get('setup_multiplier') or 1) or 1
    dispatch_label = form.get('dispatch_label') or '不派工'
    dispatch_fee = round(safe_number(dispatch_map().get(dispatch_label, 0)), 0)
    lock_install_qty = safe_int(form.get('lock_install_qty') or 0)
    lock_install_unit_price = round(safe_number(form.get('lock_install_unit_price') or 3800), 0)
    lock_install_fee = round(lock_install_qty * lock_install_unit_price, 0)
    curtain_rows = parse_curtain_rows(form)
    curtain_type = curtain_rows[0].get('type','') if curtain_rows else ''
    curtain_motor_price = curtain_rows[0].get('motor_price',0) if curtain_rows else 0
    curtain_track_unit_price = curtain_rows[0].get('track_unit_price',0) if curtain_rows else 0
    curtain_track_length = curtain_rows[0].get('track_length',0) if curtain_rows else 0
    curtain_fabric_width = curtain_rows[0].get('width',0) if curtain_rows else 0
    curtain_fabric_height = curtain_rows[0].get('height',0) if curtain_rows else 0
    curtain_fabric_unit_price = curtain_rows[0].get('cloth_price',0) if curtain_rows else 0
    curtain_install_qty = safe_int(form.get('curtain_install_qty') or sum(r.get('qty',0) for r in curtain_rows) or 0)
    curtain_install_unit = '組' if curtain_install_qty else ''
    curtain_install_amount = round(safe_number(form.get('curtain_install_amount') or 1500), 0)
    curtain_install_total = round(curtain_install_qty * curtain_install_amount, 0)
    curtain_note = curtain_rows[0].get('note','') if curtain_rows else ''
    weak_current_qty = safe_int(form.get('weak_current_qty') or 0)
    weak_current_unit = form.get('weak_current_unit') or '式'
    weak_current_amount = round(safe_number(form.get('weak_current_amount') or 0) * weak_current_qty, 0)
    hardware_qty = safe_int(form.get('hardware_qty') or 0)
    hardware_unit = form.get('hardware_unit') or '式'
    hardware_amount = round(safe_number(form.get('hardware_amount') or 0) * hardware_qty, 0)
    water_elec_qty = safe_int(form.get('water_elec_qty') or 0)
    water_elec_unit = form.get('water_elec_unit') or '式'
    water_elec_amount = round(safe_number(form.get('water_elec_amount') or 0) * water_elec_qty, 0)
    item_notes = form.getlist('item_notes')
    custom_fee_rows = parse_custom_fee_rows(form)

    items = []
    product_subtotal = 0.0
    total_cost = 0.0
    planning_fee_total = 0.0
    setup_fee_total = 0.0
    selected_products = []

    for idx, (product_id, qty) in enumerate(zip(product_ids, qtys)):
        if not product_id:
            continue
        qty = int(safe_number(qty))
        if qty <= 0:
            continue
        product = product_map.get(int(product_id))
        if not product:
            continue
        selected_products.append(product)
        row_note = (item_notes[idx] if idx < len(item_notes) else '') or product.note or ''
        unit_price = round(get_price_by_level(product, selected_level, discount_map), 0)
        line_total = round(unit_price * qty, 0)
        line_cost = round(safe_number(product.final_cost_twd) * qty, 0)
        line_planning = round(safe_number(product.planning_fee_twd) * planning_multiplier * qty, 0)
        line_setup = round(safe_number(product.setup_fee_twd) * setup_multiplier * qty, 0)
        product_subtotal += line_total
        total_cost += line_cost
        planning_fee_total += line_planning
        setup_fee_total += line_setup
        items.append({
            "product": product,
            "qty": qty,
            "unit_price": unit_price,
            "line_total": line_total,
            "cost_total": line_cost,
            "planning_fee": line_planning,
            "setup_fee": line_setup,
            "image": get_image_src(product),
            "note": row_note,
        })

    custom_fee_total = round(sum(round(safe_number(r.get("total") or 0), 0) for r in custom_fee_rows), 0)
    extra_fee_total = round(lock_install_fee + weak_current_amount + hardware_amount + water_elec_amount + custom_fee_total, 0)
    subtotal = round(product_subtotal + planning_fee_total + setup_fee_total + dispatch_fee + round(sum(r.get('line_total',0) for r in curtain_rows),0) + curtain_install_total + extra_fee_total, 0)
    tax_amount = round(subtotal * 0.05, 0)
    total_amount = round(subtotal + tax_amount, 0)
    negotiated_discount_pct = safe_number(form.get('negotiated_discount_pct') or 0)
    negotiated_total = round(safe_number(form.get('negotiated_total') or 0), 0)
    if not negotiated_total and negotiated_discount_pct > 0:
        negotiated_total = round(total_amount * (negotiated_discount_pct / 100.0), 0)
    final_for_payment = negotiated_total if negotiated_total > 0 else total_amount
    profit_amount = round(subtotal - total_cost, 0)
    gross_profit_rate = round((profit_amount / subtotal * 100.0), 1) if subtotal > 0 else 0.0
    if final_for_payment > 150000:
        payment_scheme = '30/40/30'
        deposit_1 = round(final_for_payment * 0.30, 0)
        deposit_2 = round(final_for_payment * 0.40, 0)
        deposit_3 = round(final_for_payment * 0.30, 0)
    elif final_for_payment > 50000:
        payment_scheme = '60/40'
        deposit_1 = round(final_for_payment * 0.60, 0)
        deposit_2 = 0.0
        deposit_3 = round(final_for_payment * 0.40, 0)
    else:
        payment_scheme = '一次付清'
        deposit_1 = round(final_for_payment, 0)
        deposit_2 = 0.0
        deposit_3 = 0.0

    default_note = build_auto_note(selected_products)
    note = form.get('note', '').strip() or default_note

    return {
        "dealer": dealer,
        "price_level": selected_level,
        "planning_multiplier": planning_multiplier,
        "setup_multiplier": setup_multiplier,
        "dispatch_label": dispatch_label,
        "dispatch_fee": dispatch_fee,
        "lock_install_qty": lock_install_qty,
        "lock_install_unit_price": lock_install_unit_price,
        "lock_install_fee": lock_install_fee,
        "curtain_type": curtain_type,
        "curtain_motor_price": curtain_motor_price,
        "curtain_track_unit_price": curtain_track_unit_price,
        "curtain_track_length": curtain_track_length,
        "curtain_fabric_width": curtain_fabric_width,
        "curtain_fabric_height": curtain_fabric_height,
        "curtain_fabric_unit_price": curtain_fabric_unit_price,
        "curtain_note": curtain_note,
        "curtain_rows": curtain_rows,
        "curtain_install_qty": curtain_install_qty,
        "curtain_install_unit": curtain_install_unit,
        "curtain_install_amount": curtain_install_amount,
        "curtain_install_total": curtain_install_total,
        "weak_current_qty": weak_current_qty,
        "weak_current_unit": weak_current_unit,
        "weak_current_amount": weak_current_amount,
        "hardware_qty": hardware_qty,
        "hardware_unit": hardware_unit,
        "hardware_amount": hardware_amount,
        "water_elec_qty": water_elec_qty,
        "water_elec_unit": water_elec_unit,
        "water_elec_amount": water_elec_amount,
        "custom_fee_rows": custom_fee_rows,
        "custom_fee_total": custom_fee_total,
        "extra_fee_total": extra_fee_total,
        "items": items,
        "product_subtotal": round(product_subtotal, 0),
        "planning_fee_total": round(planning_fee_total, 0),
        "setup_fee_total": round(setup_fee_total, 0),
        "subtotal": subtotal,
        "tax_amount": tax_amount,
        "total_amount": total_amount,
        "negotiated_discount_pct": negotiated_discount_pct,
        "negotiated_total": negotiated_total,
        "final_for_payment": final_for_payment,
        "total_cost": round(total_cost, 0),
        "profit_amount": profit_amount,
        "gross_profit_rate": gross_profit_rate,
        "payment_scheme": payment_scheme,
        "deposit_1": deposit_1,
        "deposit_2": deposit_2,
        "deposit_3": deposit_3,
        "note": note,
        "default_note": default_note,
    }


@flask_app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_DIR, filename)


@flask_app.route('/')
def home():
    if is_internal_logged_in():
        return redirect(url_for('internal_home'))
    if session.get('dealer_id'):
        return redirect(url_for('dealer_products'))
    return render_template('portal_landing.html')


@flask_app.route('/login', methods=['GET', 'POST'])
def internal_login():
    error = ''
    username_value = (request.form.get('username') or '').strip()
    if request.method == 'POST':
        db = db_session()
        try:
            ensure_default_internal_users(db)
            username = username_value
            password = normalize_password_input(request.form.get('password', ''))
            user = db.query(models.InternalUser).filter(models.InternalUser.username == username).first()
            if not user or not user.is_active:
                error = '帳號不存在或已停用。'
            elif normalize_password_input(user.password or '') != password:
                error = '帳號或密碼錯誤。'
            else:
                session.clear()
                session['internal_auth'] = True
                session['internal_user_id'] = user.id
                session['internal_username'] = user.username
                session['internal_display_name'] = user.display_name or user.username
                session['internal_role'] = user.role or 'quote_only'
                next_url = request.args.get('next') or url_for('internal_home')
                return redirect(next_url)
        finally:
            db.close()
    return render_template('internal_login.html', error=error, username_value=username_value)


@flask_app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    mode = (request.args.get('mode') or request.form.get('mode') or 'internal').strip()
    if mode not in ['internal', 'dealer']:
        mode = 'internal'
    message = ''
    error = ''
    username_value = (request.form.get('username') or '').strip()
    tax_id_value = normalize_tax_id_input(request.form.get('tax_id', '')) if request.method == 'POST' else ''
    db = db_session()
    try:
        if request.method == 'POST':
            if mode == 'internal':
                username = username_value
                user = db.query(models.InternalUser).filter(models.InternalUser.username == username).first() if username else None
                if not user:
                    error = '找不到這個內部帳號。'
                else:
                    ok, msg = create_password_reset_request(db, 'internal', user.username, user.display_name or user.username)
                    if ok:
                        message = msg
                        username_value = ''
                    else:
                        error = msg
            else:
                tax_id = tax_id_value
                dealer = None
                if tax_id:
                    dealer_rows = db.query(models.Dealer).filter(models.Dealer.tax_id.isnot(None)).all()
                    dealer = next((d for d in dealer_rows if normalize_tax_id_input(d.tax_id or '') == tax_id), None)
                if not dealer:
                    error = '找不到這個經銷商帳號。'
                else:
                    ok, msg = create_password_reset_request(db, 'dealer', normalize_tax_id_input(dealer.tax_id or ''), dealer.name)
                    if ok:
                        message = msg
                        tax_id_value = ''
                    else:
                        error = msg
        return render_template('forgot_password.html', mode=mode, message=message, error=error, username_value=username_value, tax_id_value=tax_id_value)
    finally:
        db.close()


@flask_app.route('/admin/reset-requests', methods=['GET', 'POST'])
@require_internal_permission('all')
def admin_reset_requests_page():
    db = db_session()
    try:
        message = None
        if request.method == 'POST':
            request_id = int(request.form.get('request_id') or 0)
            action_type = (request.form.get('action_type') or 'reset').strip()
            row = db.query(models.PasswordResetRequest).filter(models.PasswordResetRequest.id == request_id).first()
            if row and row.status == 'pending':
                admin_name = session.get('internal_display_name') or session.get('internal_username') or DEFAULT_SUPERVISOR_USERNAME
                if action_type == 'reject':
                    row.status = 'rejected'
                    row.admin_note = (request.form.get('admin_note') or '').strip()
                    row.resolved_by = admin_name
                    row.resolved_at = datetime.utcnow()
                    db.commit()
                    message = '已駁回忘記密碼申請。'
                else:
                    new_password = normalize_password_input(request.form.get('new_password') or '')
                    if not new_password:
                        message = '請輸入新密碼。'
                    elif row.account_type == 'internal':
                        user = db.query(models.InternalUser).filter(models.InternalUser.username == row.account_identifier).first()
                        if user:
                            user.password = new_password
                            user.updated_at = datetime.utcnow()
                            row.status = 'reset'
                            row.admin_note = (request.form.get('admin_note') or '').strip()
                            row.resolved_by = admin_name
                            row.resolved_at = datetime.utcnow()
                            db.commit()
                            message = f'已重設內部帳號：{user.username}'
                    elif row.account_type == 'dealer':
                        dealer_rows = db.query(models.Dealer).filter(models.Dealer.tax_id.isnot(None)).all()
                        dealer = next((d for d in dealer_rows if normalize_tax_id_input(d.tax_id or '') == row.account_identifier), None)
                        if dealer:
                            dealer.access_key = new_password
                            row.status = 'reset'
                            row.admin_note = (request.form.get('admin_note') or '').strip()
                            row.resolved_by = admin_name
                            row.resolved_at = datetime.utcnow()
                            db.commit()
                            message = f'已重設經銷商帳號：{dealer.name}'
        pending_rows = db.query(models.PasswordResetRequest).filter(models.PasswordResetRequest.status == 'pending').order_by(models.PasswordResetRequest.created_at.desc()).all()
        history_rows = db.query(models.PasswordResetRequest).filter(models.PasswordResetRequest.status != 'pending').order_by(models.PasswordResetRequest.created_at.desc()).limit(50).all()
        return render_template('reset_requests_admin.html', pending_rows=pending_rows, history_rows=history_rows, message=message)
    finally:
        db.close()


@flask_app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('home'))


@flask_app.route('/dealer/login', methods=['GET', 'POST'])
def dealer_login():
    db = db_session()
    error = ''
    tax_id_value = normalize_tax_id_input(request.form.get('tax_id', '')) if request.method == 'POST' else ''
    try:
        if request.method == 'POST':
            tax_id = tax_id_value
            access_key = normalize_password_input(request.form.get('access_key', ''))
            dealer = None
            if tax_id:
                dealer_rows = db.query(models.Dealer).filter(models.Dealer.tax_id.isnot(None)).all()
                dealer = next((d for d in dealer_rows if normalize_tax_id_input(d.tax_id or '') == tax_id), None)
            if dealer and normalize_password_input(dealer.access_key or '0000') == access_key:
                session.clear()
                session['dealer_id'] = dealer.id
                next_url = request.args.get('next') or url_for('dealer_products')
                return redirect(next_url)
            error = '統編或密碼錯誤。'
        return render_template('dealer_login.html', error=error, tax_id_value=tax_id_value)
    finally:
        db.close()


@flask_app.route('/dealer/logout')
def dealer_logout():
    session.clear()
    return redirect(url_for('home'))


@flask_app.route('/admin/password', methods=['GET', 'POST'])
@require_internal_login
def admin_password_page():
    if request.method == 'POST':
        current_password = normalize_password_input(request.form.get('current_password', ''))
        new_password = normalize_password_input(request.form.get('new_password', ''))
        confirm_password = normalize_password_input(request.form.get('confirm_password', ''))
        db = db_session()
        try:
            user = current_internal_user(db)
            if user and current_password == normalize_password_input(user.password or '') and new_password and len(new_password) <= 10 and new_password == confirm_password:
                user.password = new_password
                user.updated_at = datetime.utcnow()
                db.commit()
                session.clear()
                return redirect(url_for('internal_login'))
            flash('目前密碼不正確，或新密碼設定有誤。')
        finally:
            db.close()
        return redirect(url_for('admin_password_page'))
    return render_template('password_form.html', mode='internal', title_text='修改內部密碼', description='最多 10 碼，儲存後會重新登入。')


@flask_app.route('/dealer/password', methods=['GET', 'POST'])
@require_dealer_login
def dealer_password_page():
    db = db_session()
    try:
        dealer = current_dealer(db)
        if not dealer:
            session.clear()
            return redirect(url_for('dealer_login'))
        if request.method == 'POST':
            current_password = normalize_password_input(request.form.get('current_password', ''))
            new_password = normalize_password_input(request.form.get('new_password', ''))
            confirm_password = normalize_password_input(request.form.get('confirm_password', ''))
            if current_password == normalize_password_input(dealer.access_key or '0000') and new_password and len(new_password) <= 10 and new_password == confirm_password:
                dealer.access_key = new_password
                db.commit()
                session.clear()
                return redirect(url_for('dealer_login'))
            return redirect(url_for('dealer_password_page'))
        return render_template('password_form.html', mode='dealer', title_text='修改經銷商密碼', description='最多 10 碼，儲存後會重新登入。')
    finally:
        db.close()


@flask_app.route('/internal/home')
@require_internal_login
def internal_home():
    if not (session.get('internal_role') == 'supervisor' or session.get('internal_auth')):
        return redirect('/quotes/new')
    q = request.args.get('q', '')
    category = request.args.get('category', '')
    level = request.args.get('level', '市場報價')
    selected_extra_cols = request.args.getlist('extra_cols')

    db = db_session()
    try:
        products_query = db.query(models.Product)
        if q:
            keyword = f"%{q}%"
            products_query = products_query.filter(
                or_(
                    models.Product.model.like(keyword),
                    models.Product.name.like(keyword),
                    models.Product.category.like(keyword),
                )
            )
        if category:
            products_query = products_query.filter(models.Product.category == category)
        products = products_query.order_by(models.Product.category, models.Product.model).all()
        categories = [x[0] for x in db.query(models.Product.category).distinct().all() if x[0]]
        discount_map = get_price_setting_map(db)
        rates = get_rate_map(db)
        extra_labels = {'cost':'成本','market_min':'市場最低價','designer':'設計師價','dealer_lv1':'一級經銷商','dealer_lv2':'二級經銷商','branch':'分公司','master':'總經銷商','profit':'利潤率'}
        selected_extra_labels = '、'.join([extra_labels[k] for k in selected_extra_cols if k in extra_labels]) if selected_extra_cols else '請勾選要顯示的欄位'
        return render_template(
            'home.html',
            products=products,
            categories=categories,
            selected_category=category,
            keyword=q,
            selected_level=level,
            selected_extra_cols=selected_extra_cols,
            discount_map=discount_map,
            get_price_by_level=get_price_by_level,
            get_profit_rate=get_profit_rate,
            get_image_src=get_image_src,
            price_levels=PRICE_LEVELS,
            extra_choices=[
                {"value": "cost", "label": "成本"},
                {"value": "market_min", "label": "市場最低價"},
                {"value": "designer", "label": "設計師價"},
                {"value": "dealer_lv1", "label": "一級經銷商"},
                {"value": "dealer_lv2", "label": "二級經銷商"},
                {"value": "branch", "label": "分公司"},
                {"value": "master", "label": "總經銷商"},
                {"value": "profit", "label": "利潤率"},
            ],
            profit_level=(
                '市場報價' if not selected_extra_cols else (
                    {"market_min": "市場最低價", "designer": "設計師價", "dealer_lv1": "一級經銷商", "dealer_lv2": "二級經銷商", "branch": "分公司", "master": "總經銷商"}.get(selected_extra_cols[0])
                    if len(selected_extra_cols) == 1 and selected_extra_cols[0] not in ['cost','profit'] else ('市場報價' if selected_extra_cols==['profit'] else None)
                )
            ),
            extra_labels=extra_labels,
            selected_extra_labels=selected_extra_labels,
            usd_rate=rates['USD'],
        )
    finally:
        db.close()


DEALER_LEVEL_ORDER = ['分公司', '總經銷商', '一級經銷商', '二級經銷商', '設計師價', '市場最低價', '市場報價']
DEALER_LEVEL_LABEL_TO_FIELD = {
    '分公司': 'branch_price',
    '總經銷商': 'master_dealer_price',
    '一級經銷商': 'dealer_lv1_price',
    '二級經銷商': 'dealer_lv2_price',
    '設計師價': 'designer_price',
    '市場最低價': 'market_min_price',
    '市場報價': 'market_price',
}


def _dealer_price_field(level):
    return DEALER_LEVEL_LABEL_TO_FIELD.get(level, 'market_price')


def _allowed_levels_for_dealer(level):
    level = (level or '市場報價').strip()
    if level not in DEALER_LEVEL_ORDER:
        return ['市場報價']
    start = DEALER_LEVEL_ORDER.index(level)
    return DEALER_LEVEL_ORDER[start:]


@flask_app.route('/dealer/products')
@require_dealer_login
def dealer_products():
    q = request.args.get('q', '')
    category = request.args.get('category', '')
    status = request.args.get('status', '')
    selected_export_levels = request.args.getlist('export_levels')
    include_image = bool(request.args.get('include_image'))
    include_description = bool(request.args.get('include_description', '1'))
    db = db_session()
    try:
        dealer = current_dealer(db)
        if not dealer_permission_enabled(dealer, 'can_view_products'):
            flash('此經銷商目前未開啟查價權限。')
            return redirect(url_for('dealer_logout'))
        products_query = db.query(models.Product)
        if q:
            keyword = f"%{q}%"
            products_query = products_query.filter(or_(models.Product.model.like(keyword), models.Product.name.like(keyword), models.Product.category.like(keyword)))
        if category:
            products_query = products_query.filter(models.Product.category == category)
        if status:
            products_query = products_query.filter(models.Product.status == status)
        products = products_query.order_by(models.Product.category, models.Product.model).all()
        categories = [x[0] for x in db.query(models.Product.category).distinct().all() if x[0]]
        statuses = [x[0] for x in db.query(models.Product.status).distinct().all() if x[0]]
        price_level = dealer.level or '市場報價'
        allowed_export_levels = _allowed_levels_for_dealer(price_level)
        if not selected_export_levels:
            selected_export_levels = [price_level]
        selected_export_levels = [x for x in selected_export_levels if x in allowed_export_levels]
        return render_template(
            'dealer_products.html', dealer=dealer, products=products, categories=categories, statuses=statuses,
            keyword=q, selected_category=category, selected_status=status, price_level=price_level,
            get_image_src=get_image_src, get_price_by_level=get_price_by_level, discount_map=get_price_setting_map(db),
            can_export=dealer_permission_enabled(dealer, 'can_export_prices'), can_quote=dealer_permission_enabled(dealer, 'can_create_quote'),
            allowed_export_levels=allowed_export_levels, selected_export_levels=selected_export_levels,
            include_image=include_image, include_description=include_description,
        )
    finally:
        db.close()


@flask_app.route('/dealer/export')
@require_dealer_login
def dealer_export():
    db = db_session()
    try:
        dealer = current_dealer(db)
        if not dealer_permission_enabled(dealer, 'can_export_prices'):
            flash('此經銷商目前未開啟匯出權限。')
            return redirect(url_for('dealer_products'))
        q = request.args.get('q', '')
        category = request.args.get('category', '')
        status = request.args.get('status', '')
        include_image = bool(request.args.get('include_image'))
        include_description = bool(request.args.get('include_description', '1'))
        include_category = bool(request.args.get('include_category', '1'))
        include_status = bool(request.args.get('include_status'))
        selected_levels = request.args.getlist('export_levels')
        selected_ids = [int(x) for x in request.args.getlist('selected_product_ids') if str(x).isdigit()]
        allowed_levels = _allowed_levels_for_dealer(dealer.level or '市場報價')
        selected_levels = [x for x in selected_levels if x in allowed_levels] or [dealer.level or '市場報價']
        products_query = db.query(models.Product)
        if q:
            keyword = f"%{q}%"
            products_query = products_query.filter(or_(models.Product.model.like(keyword), models.Product.name.like(keyword), models.Product.category.like(keyword)))
        if category:
            products_query = products_query.filter(models.Product.category == category)
        if status:
            products_query = products_query.filter(models.Product.status == status)
        if selected_ids:
            products_query = products_query.filter(models.Product.id.in_(selected_ids))
        else:
            flash('請先勾選要匯出的商品。')
            return redirect(url_for('dealer_products', q=q, category=category, status=status, include_image='1' if include_image else '', include_description='1' if include_description else ''))
        products = products_query.order_by(models.Product.category, models.Product.model).all()
        wb = Workbook()
        ws = wb.active
        ws.title = '價格表'
        headers = []
        if include_category:
            headers.append('類別')
        if include_status:
            headers.append('狀態')
        headers.extend(['型號','品名'])
        if include_description:
            headers.append('產品描述')
        if include_image:
            headers.append('圖片')
        headers.append('單位')
        headers.extend(selected_levels)
        headers.append('備註')
        ws.append(headers)
        image_col = headers.index('圖片') + 1 if include_image else None
        row_index = 2
        discount_map = get_price_setting_map(db)
        for p in products:
            row = []
            if include_category:
                row.append(p.category or '')
            if include_status:
                row.append(p.status or '')
            row.extend([p.model, p.name])
            if include_description:
                row.append(p.description or '')
            if include_image:
                row.append('')
            row.append(p.unit or '')
            for level in selected_levels:
                row.append(round(get_price_by_level(p, level, discount_map) or 0))
            row.append(p.note or '')
            ws.append(row)
            ws.row_dimensions[row_index].height = 48
            if include_image:
                img_file, _ = resolve_product_image_file(p)
                if img_file and img_file.exists():
                    try:
                        xl_img = XLImage(str(img_file))
                        xl_img.width = 46
                        xl_img.height = 46
                        ws.add_image(xl_img, f"{chr(64+image_col)}{row_index}")
                    except Exception:
                        pass
            row_index += 1
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        filename = f'{safe_filename(dealer.name)}_價格表.xlsx'
        log_event(db, 'export', 'dealer_products', f'經銷商匯出價格表：{dealer.name}/{"/".join(selected_levels)}', str(dealer.id))
        return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    finally:
        db.close()


@flask_app.route('/products')
@require_internal_permission('all')
def products_page():
    db = db_session()
    try:
        products = db.query(models.Product).order_by(models.Product.category, models.Product.model).all()
        discount_map = get_price_setting_map(db)
        rates = get_rate_map(db)
        for p in products:
            p.usd_rate = rates['USD']
            p.rmb_rate = rates.get('RMB', 4.5)
        export_options = [
            ('market_price', '市場報價'),
            ('market_min_price', '市場最低價'),
            ('designer_price', '設計師價'),
            ('dealer_lv1_price', '一級經銷商'),
            ('dealer_lv2_price', '二級經銷商'),
            ('branch_price', '分公司'),
            ('master_dealer_price', '總經銷商'),
        ]
        categories, statuses = get_option_lists(db)
        return render_template('products.html', products=products, get_image_src=get_image_src, discount_map=discount_map, get_price_by_level=get_price_by_level, get_profit_rate=get_profit_rate, export_options=export_options, usd_rate=rates['USD'], categories=categories, statuses=statuses)
    finally:
        db.close()




@flask_app.route('/products/export', methods=['POST'])
@require_internal_permission('all')
def products_export():
    db = db_session()
    try:
        selected_levels = request.form.getlist('export_levels')
        include_fields = {
            'description': bool(request.form.get('include_description')),
            'image': bool(request.form.get('include_image')),
            'category': bool(request.form.get('include_category')),
            'status': bool(request.form.get('include_status')),
            'cost_twd': bool(request.form.get('include_cost_twd')),
            'cost_usd': bool(request.form.get('include_cost_usd')),
            'special_ratio': bool(request.form.get('include_special_ratio')),
            'outsourced_fee': bool(request.form.get('include_outsourced_fee')),
            'duty_rate_pct': bool(request.form.get('include_duty_rate_pct')),
            'duty_cost_usd': bool(request.form.get('include_duty_cost_usd')),
            'planning_fee': bool(request.form.get('include_planning_fee')),
            'setup_fee': bool(request.form.get('include_setup_fee')),
            'profit_rate': bool(request.form.get('include_profit_rate')),
        }
        selected_ids = [int(x) for x in request.form.getlist('selected_product_ids') if str(x).isdigit()]
        products_query = db.query(models.Product)
        if selected_ids:
            products_query = products_query.filter(models.Product.id.in_(selected_ids))
        else:
            flash('請先勾選要匯出的商品。')
            return redirect(url_for('products_page'))
        products = products_query.order_by(models.Product.category, models.Product.model).all()
        rates = get_rate_map(db)
        for p in products:
            p.usd_rate = rates['USD']
            p.rmb_rate = rates.get('RMB', 4.5)
        bio = build_product_export_workbook(products, selected_levels, include_fields)
        file_tag = '_'.join(selected_export_labels(selected_levels)) or '市場報價'
        filename = f'商品總表_{file_tag}.xlsx'
        log_event(db, 'export', 'products', f'匯出商品總表：{file_tag}')
        return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    finally:
        db.close()



def _current_products_with_rates(db):
    products = db.query(models.Product).order_by(models.Product.category, models.Product.model).all()
    rates = get_rate_map(db)
    for p in products:
        p.usd_rate = rates['USD']
    return products, rates


def _render_products_pdf(products, include_fields):
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.pdfgen import canvas
    pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
    bio = BytesIO()
    c = canvas.Canvas(bio, pagesize=landscape(A4))
    width, height = landscape(A4)
    c.setFont('STSong-Light', 14)
    c.drawString(30, height-30, '商品總表')
    y = height - 55
    headers = ['型號','品名']
    if include_fields.get('description'): headers.append('描述')
    if include_fields.get('cost_twd'): headers.append('成本TWD')
    if include_fields.get('cost_usd'): headers.append('成本USD')
    headers.append('市場報價')
    c.setFont('STSong-Light', 8)
    x_positions = [30,120,300,520,610,700]
    for i,h in enumerate(headers):
        c.drawString(x_positions[i], y, h)
    y -= 14
    for p in products[:120]:
        vals=[p.model, p.name]
        if include_fields.get('description'): vals.append((p.description or '')[:28])
        if include_fields.get('cost_twd'): vals.append(f"{round(safe_number(p.final_cost_twd),0):,.0f}")
        if include_fields.get('cost_usd'): vals.append(f"US$ {((safe_number(p.final_cost_twd)/(p.usd_rate or 32.5)) if (p.usd_rate or 32.5) else 0):.2f}")
        vals.append(f"{round(safe_number(p.market_price),0):,.0f}")
        for i,v in enumerate(vals):
            c.drawString(x_positions[i], y, str(v))
        y -= 12
        if y < 40:
            c.showPage(); c.setFont('STSong-Light',8); y = height - 40
    c.save(); bio.seek(0); return bio

@flask_app.route('/products/export/pdf')
@require_internal_permission('all')
def products_export_pdf():
    db = db_session()
    try:
        products, rates = _current_products_with_rates(db)
        include_fields = {'description': True, 'image': False, 'cost_twd': True, 'cost_usd': True, 'planning_fee': False, 'setup_fee': False, 'profit_rate': False}
        bio = _render_products_pdf(products, include_fields)
        log_event(db, 'export', 'products', '匯出商品總表PDF')
        return send_file(bio, as_attachment=True, download_name='商品總表.pdf', mimetype='application/pdf')
    finally:
        db.close()

@flask_app.route('/products/print')
def products_print_page():
    db = db_session()
    try:
        products, rates = _current_products_with_rates(db)
        return render_template('products_print.html', products=products, usd_rate=rates['USD'])
    finally:
        db.close()

@flask_app.route('/admin/import', methods=['GET', 'POST'])
@require_internal_permission('all')
def admin_import_page():
    db = db_session()
    try:
        message = None
        if request.method == 'POST':
            action_type = request.form.get('action_type', 'import_excel')
            if action_type == 'import_excel':
                file = request.files.get('file')
                if file and file.filename:
                    save_path = UPLOAD_DIR / file.filename
                    file.save(save_path)
                    result = import_all(db, str(save_path))
                    message = f"匯入完成：商品 {result['products']} 筆，經銷商 {result['dealers']} 筆。"
                    log_event(db, 'import', 'excel', message)
                else:
                    message = '請先選擇 Excel 檔案。'
            elif action_type == 'add_category':
                name = (request.form.get('category_name') or '').strip()
                if name:
                    exists = db.query(models.CategoryOption).filter(models.CategoryOption.name == name).first()
                    if not exists:
                        db.add(models.CategoryOption(name=name, is_active=1))
                        db.commit()
                        message = f'已新增類別：{name}'
                        log_event(db, 'update', 'category_option', message)
                    else:
                        message = f'類別已存在：{name}'
                else:
                    message = '請輸入類別名稱。'
            elif action_type == 'delete_category':
                cid = request.form.get('category_id')
                item = db.query(models.CategoryOption).filter(models.CategoryOption.id == int(cid or 0)).first()
                if item:
                    name = item.name
                    db.delete(item)
                    db.commit()
                    message = f'已刪除類別：{name}'
                    log_event(db, 'delete', 'category_option', message)
            elif action_type == 'add_status':
                name = (request.form.get('status_name') or '').strip()
                if name:
                    exists = db.query(models.StatusOption).filter(models.StatusOption.name == name).first()
                    if not exists:
                        db.add(models.StatusOption(name=name, is_active=1))
                        db.commit()
                        message = f'已新增狀態：{name}'
                        log_event(db, 'update', 'status_option', message)
                    else:
                        message = f'狀態已存在：{name}'
                else:
                    message = '請輸入狀態名稱。'
            elif action_type == 'delete_status':
                sid = request.form.get('status_id')
                item = db.query(models.StatusOption).filter(models.StatusOption.id == int(sid or 0)).first()
                if item:
                    name = item.name
                    db.delete(item)
                    db.commit()
                    message = f'已刪除狀態：{name}'
                    log_event(db, 'delete', 'status_option', message)
        rates = get_rate_map(db)
        category_options = db.query(models.CategoryOption).order_by(models.CategoryOption.name).all()
        status_options = db.query(models.StatusOption).order_by(models.StatusOption.name).all()
        return render_template('admin_import.html', rates=rates, message=message, category_options=category_options, status_options=status_options)
    finally:
        db.close()


@flask_app.route('/admin/rates', methods=['GET', 'POST'])
@require_internal_permission('all')
def admin_rates_page():
    db = db_session()
    try:
        message = None
        if request.method == 'POST':
            usd_rate = float(request.form.get('usd_rate', 32.5) or 32.5)
            rmb_rate = float(request.form.get('rmb_rate', 4.5) or 4.5)
            ensure_default_rates(db, usd_rate=usd_rate, rmb_rate=rmb_rate)
            logo_file = request.files.get('company_logo')
            if logo_file and logo_file.filename:
                ext = Path(logo_file.filename).suffix.lower() or '.png'
                save_name = f'company_logo{ext}'
                logo_path = UPLOAD_DIR / save_name
                for old in ['company_logo.png','company_logo.jpg','company_logo.jpeg']:
                    oldp = UPLOAD_DIR / old
                    if oldp.exists() and oldp.name != save_name:
                        oldp.unlink()
                logo_file.save(logo_path)
            if request.form.get('apply_to_products'):
                apply_discount_settings_to_products(db)
                message = '匯率已更新，並已重算全部商品售價。'
            else:
                message = '匯率已更新。'
            log_event(db, 'update', 'rates', f'USD={usd_rate}, RMB={rmb_rate}, 匯率已更新')
        rates = get_rate_map(db)
        discount_map = get_product_discount_display_map(db)
        average_discount_map = average_product_level_ratios(db)
        return render_template('admin_rates.html', rates=rates, discount_map=discount_map, average_discount_map=average_discount_map, message=message, company_logo_src=get_company_logo_src())
    finally:
        db.close()



@flask_app.route('/admin/products/new', methods=['GET', 'POST'])
@require_internal_permission('all')
def admin_product_new():
    db = db_session()
    try:
        if request.method == 'POST':
            category = (request.form.get('category') or '').strip()
            model = request.form.get('model', '').strip()
            name = request.form.get('name', '').strip()
            if not model or not name:
                categories, statuses = get_option_lists(db)
                return render_template('product_form.html', product=None, price_levels=PRICE_LEVELS, action='/admin/products/new', title_text='新增商品', error='型號與品名必填', categories=categories, statuses=statuses, change_logs=[], discount_ratio_map=get_product_discount_display_map(db), product_discount_ratio_map=get_product_discount_display_map(db))
            if db.query(models.Product).filter(models.Product.model == model).first():
                categories, statuses = get_option_lists(db)
                return render_template('product_form.html', product=None, price_levels=PRICE_LEVELS, action='/admin/products/new', title_text='新增商品', error='型號已存在', categories=categories, statuses=statuses, change_logs=[], discount_ratio_map=get_product_discount_display_map(db), product_discount_ratio_map=get_product_discount_display_map(db))
            rates = get_rate_map(db)
            discount_map = get_price_setting_map(db)
            product = models.Product(
                category=category,
                model=model,
                name=name,
                description=request.form.get('description', ''),
                note=request.form.get('note', ''),
                unit=request.form.get('unit', '台'),
                status=(request.form.get('status') or '').strip(),
                source_currency=request.form.get('source_currency', 'TWD'),
                source_cost=float(request.form.get('source_cost', 0) or 0),
                shipping_usd=float(request.form.get('shipping_usd', 0) or 0),
                duty_rate_pct=normalize_percentage_value(request.form.get('duty_rate_pct', 0) or 0),
                outsourced_parts_fee_twd=round(float(request.form.get('outsourced_parts_fee_twd', 0) or 0),0),
                planning_fee_twd=round(float(request.form.get('planning_fee_twd', 0) or 0),0),
                setup_fee_twd=round(float(request.form.get('setup_fee_twd', 0) or 0),0),
                special_market_price_ratio=round3(request.form.get('special_market_price_ratio', 1), 1),
                special_market_min_ratio=round3(request.form.get('special_market_min_ratio', 1), 1),
                special_designer_ratio=round3(request.form.get('special_designer_ratio', 1), 1),
                special_dealer_lv1_ratio=round3(request.form.get('special_dealer_lv1_ratio', 1), 1),
                special_dealer_lv2_ratio=round3(request.form.get('special_dealer_lv2_ratio', 1), 1),
                special_branch_ratio=round3(request.form.get('special_branch_ratio', 1), 1),
                special_master_ratio=round3(request.form.get('special_master_ratio', 1), 1),
                market_min_discount_ratio=normalize_percentage_value(request.form.get('market_min_discount_ratio', discount_map.get('市場最低價', 95)) or discount_map.get('市場最低價', 95)),
                designer_discount_ratio=normalize_percentage_value(request.form.get('designer_discount_ratio', discount_map.get('設計師價', 90)) or discount_map.get('設計師價', 90)),
                dealer_lv1_discount_ratio=normalize_percentage_value(request.form.get('dealer_lv1_discount_ratio', discount_map.get('一級經銷商', 85)) or discount_map.get('一級經銷商', 85)),
                dealer_lv2_discount_ratio=normalize_percentage_value(request.form.get('dealer_lv2_discount_ratio', discount_map.get('二級經銷商', 80)) or discount_map.get('二級經銷商', 80)),
                branch_discount_ratio=normalize_percentage_value(request.form.get('branch_discount_ratio', discount_map.get('分公司', 75)) or discount_map.get('分公司', 75)),
                master_discount_ratio=normalize_percentage_value(request.form.get('master_discount_ratio', discount_map.get('總經銷商', 70)) or discount_map.get('總經銷商', 70)),
                market_price=round(float(request.form.get('market_price', 0) or 0),0),
                market_min_price=round(float(request.form.get('market_min_price', 0) or 0),0),
                designer_price=round(float(request.form.get('designer_price', 0) or 0),0),
                dealer_lv1_price=round(float(request.form.get('dealer_lv1_price', 0) or 0),0),
                dealer_lv2_price=round(float(request.form.get('dealer_lv2_price', 0) or 0),0),
                branch_price=round(float(request.form.get('branch_price', 0) or 0),0),
                master_dealer_price=round(float(request.form.get('master_dealer_price', 0) or 0),0),
                image_url=request.form.get('image_url', ''),
                updated_at=datetime.utcnow(),
            )
            recompute_product_prices(product, rates, discount_map)
            db.add(product)
            db.commit()
            db.refresh(product)
            image_file = request.files.get('image_file')
            if image_file and image_file.filename:
                current_file, _ = resolve_product_image_file(product)
                if current_file and current_file.exists():
                    ensure_original_backup(product, current_file)
                target_path, rel = normalize_image_target(product, preferred_suffix='.png')
                save_square_image(image_file.stream, target_path, size=500)
                product.image_path = f'/uploads/{rel}'
                db.commit()
            log_event(db, 'create', 'product', f'商品 {product.model} 已新增', str(product.id))
            return redirect(f'/admin/products/{product.id}/edit')
        categories, statuses = get_option_lists(db)
        return render_template('product_form.html', product=None, price_levels=PRICE_LEVELS, action='/admin/products/new', title_text='新增商品', categories=categories, statuses=statuses, change_logs=[], discount_ratio_map=get_product_discount_display_map(db), product_discount_ratio_map=get_product_discount_display_map(db))
    finally:
        db.close()


@flask_app.route('/admin/products/<int:product_id>/edit', methods=['GET', 'POST'])
@require_internal_permission('all')
def admin_product_edit(product_id: int):
    db = db_session()
    try:
        product = db.query(models.Product).filter(models.Product.id == product_id).first()
        if not product:
            abort(404)
        if request.method == 'POST':
            before = {k: getattr(product, k) for k in ['category','model','name','description','note','unit','status','source_currency','source_cost','shipping_usd','duty_rate_pct','outsourced_parts_fee_twd','planning_fee_twd','setup_fee_twd','market_min_discount_ratio','designer_discount_ratio','dealer_lv1_discount_ratio','dealer_lv2_discount_ratio','branch_discount_ratio','master_discount_ratio','market_price','market_min_price','designer_price','dealer_lv1_price','dealer_lv2_price','branch_price','master_dealer_price']}
            product.category = (request.form.get('category') or '').strip()
            product.model = request.form.get('model', '').strip()
            product.name = request.form.get('name', '').strip()
            product.description = request.form.get('description', '')
            product.note = request.form.get('note', '')
            product.unit = request.form.get('unit', '台')
            product.status = (request.form.get('status') or '').strip()
            product.source_currency = request.form.get('source_currency', 'TWD')
            product.source_cost = float(request.form.get('source_cost', 0) or 0)
            product.shipping_usd = float(request.form.get('shipping_usd', 0) or 0)
            product.duty_rate_pct = normalize_percentage_value(request.form.get('duty_rate_pct', 0) or 0)
            product.outsourced_parts_fee_twd = round(float(request.form.get('outsourced_parts_fee_twd', 0) or 0),0)
            product.planning_fee_twd = round(float(request.form.get('planning_fee_twd', 0) or 0),0)
            product.setup_fee_twd = round(float(request.form.get('setup_fee_twd', 0) or 0),0)
            product.special_market_price_ratio = round3(request.form.get('special_market_price_ratio', 1), 1)
            product.special_market_min_ratio = round3(request.form.get('special_market_min_ratio', 1), 1)
            product.special_designer_ratio = round3(request.form.get('special_designer_ratio', 1), 1)
            product.special_dealer_lv1_ratio = round3(request.form.get('special_dealer_lv1_ratio', 1), 1)
            product.special_dealer_lv2_ratio = round3(request.form.get('special_dealer_lv2_ratio', 1), 1)
            product.special_branch_ratio = round3(request.form.get('special_branch_ratio', 1), 1)
            product.special_master_ratio = round3(request.form.get('special_master_ratio', 1), 1)
            discount_map = get_price_setting_map(db)
            product.market_min_discount_ratio = normalize_percentage_value(request.form.get('market_min_discount_ratio', discount_map.get('市場最低價', 95)) or discount_map.get('市場最低價', 95))
            product.designer_discount_ratio = normalize_percentage_value(request.form.get('designer_discount_ratio', discount_map.get('設計師價', 90)) or discount_map.get('設計師價', 90))
            product.dealer_lv1_discount_ratio = normalize_percentage_value(request.form.get('dealer_lv1_discount_ratio', discount_map.get('一級經銷商', 85)) or discount_map.get('一級經銷商', 85))
            product.dealer_lv2_discount_ratio = normalize_percentage_value(request.form.get('dealer_lv2_discount_ratio', discount_map.get('二級經銷商', 80)) or discount_map.get('二級經銷商', 80))
            product.branch_discount_ratio = normalize_percentage_value(request.form.get('branch_discount_ratio', discount_map.get('分公司', 75)) or discount_map.get('分公司', 75))
            product.master_discount_ratio = normalize_percentage_value(request.form.get('master_discount_ratio', discount_map.get('總經銷商', 70)) or discount_map.get('總經銷商', 70))
            rates = get_rate_map(db)
            product.market_price = round(float(request.form.get('market_price', 0) or 0),0)
            product.market_min_price = round(float(request.form.get('market_min_price', 0) or 0),0)
            product.designer_price = round(float(request.form.get('designer_price', 0) or 0),0)
            product.dealer_lv1_price = round(float(request.form.get('dealer_lv1_price', 0) or 0),0)
            product.dealer_lv2_price = round(float(request.form.get('dealer_lv2_price', 0) or 0),0)
            product.branch_price = round(float(request.form.get('branch_price', 0) or 0),0)
            product.master_dealer_price = round(float(request.form.get('master_dealer_price', 0) or 0),0)
            recompute_product_prices(product, rates, discount_map)
            product.image_url = request.form.get('image_url', '')
            image_file = request.files.get('image_file')
            if image_file and image_file.filename:
                current_file, _ = resolve_product_image_file(product)
                if current_file and current_file.exists():
                    ensure_original_backup(product, current_file)
                target_path, rel = normalize_image_target(product, preferred_suffix='.png')
                save_square_image(image_file.stream, target_path, size=500)
                product.image_path = f'/uploads/{rel}'
            product.updated_at = datetime.utcnow()
            changes=[]
            for k,v in before.items():
                nv = getattr(product,k)
                if str(v or '') != str(nv or ''):
                    changes.append(f'{k}: {v} → {nv}')
            db.commit()
            record_product_change(db, product.id, changes or ['未檢出欄位差異，但已儲存'])
            log_event(db, 'update', 'product', f'商品 {product.model} 已修改', str(product.id))
            return redirect(f'/admin/products/{product_id}/edit')
        categories, statuses = get_option_lists(db)
        change_logs = db.query(models.ProductChangeLog).filter(models.ProductChangeLog.product_id == product.id).order_by(models.ProductChangeLog.changed_at.desc()).all()
        return render_template('product_form.html', product=product, price_levels=PRICE_LEVELS, action=f'/admin/products/{product_id}/edit', title_text='修改商品', get_image_src=get_image_src, original_backup_exists=bool(get_original_backup_path(product)), categories=categories, statuses=statuses, change_logs=change_logs, discount_ratio_map=get_product_discount_display_map(db), product_discount_ratio_map=get_product_discount_ratio_map(product, get_price_setting_map(db)))
    finally:
        db.close()



@flask_app.route('/admin/products/<int:product_id>/image-upload', methods=['POST'])
@require_internal_permission('all')
def admin_product_image_upload(product_id: int):
    db = db_session()
    try:
        product = db.query(models.Product).filter(models.Product.id == product_id).first()
        if not product:
            abort(404)
        image_file = request.files.get('image_file')
        if not image_file or not image_file.filename:
            return {'ok': False, 'message': '未選擇圖片'}, 400

        target_path, rel = normalize_image_target(product, preferred_suffix='.png')
        src_file, _ = resolve_product_image_file(product)
        if src_file and src_file.exists():
            ensure_original_backup(product, src_file)
        save_square_image(image_file.stream, target_path, size=500)
        product.image_path = f'/uploads/{rel}'
        product.updated_at = datetime.utcnow()
        db.commit()
        log_event(db, 'update', 'product_image', f'商品 {product.model} 上傳/更新圖片', str(product.id))
        return {'ok': True, 'image_url': f'/uploads/{rel}?v={int(datetime.utcnow().timestamp())}'}
    finally:
        db.close()


@flask_app.route('/admin/products/<int:product_id>/image-crop-save', methods=['POST'])
@require_internal_permission('all')
def admin_product_image_crop_save(product_id: int):
    db = db_session()
    try:
        product = db.query(models.Product).filter(models.Product.id == product_id).first()
        if not product:
            abort(404)
        image_file = request.files.get('image')
        if not image_file:
            return {'ok': False, 'message': '缺少裁切圖片資料'}, 400

        current_file, _ = resolve_product_image_file(product)
        if current_file and current_file.exists():
            ensure_original_backup(product, current_file)

        target_path, rel = normalize_image_target(product, preferred_suffix='.png')
        target_path.parent.mkdir(parents=True, exist_ok=True)
        with open(target_path, 'wb') as f:
            f.write(image_file.read())

        # enforce square and consistent size
        save_square_image(target_path, target_path, size=500)
        product.image_path = f'/uploads/{rel}'
        product.updated_at = datetime.utcnow()
        db.commit()
        log_event(db, 'update', 'product_image', f'商品 {product.model} 裁切圖片', str(product.id))
        return {'ok': True, 'image_url': f'/uploads/{rel}?v={int(datetime.utcnow().timestamp())}'}
    finally:
        db.close()


@flask_app.route('/admin/products/<int:product_id>/image-reset', methods=['POST'])
@require_internal_permission('all')
def admin_product_image_reset(product_id: int):
    db = db_session()
    try:
        product = db.query(models.Product).filter(models.Product.id == product_id).first()
        if not product:
            abort(404)
        backup = get_original_backup_path(product)
        if not backup or not backup.exists():
            return {'ok': False, 'message': '找不到原始備份圖片'}, 400
        target_path, rel = normalize_image_target(product, preferred_suffix=backup.suffix.lower() or '.png')
        shutil.copy2(backup, target_path)
        save_square_image(target_path, target_path, size=500)
        product.image_path = f'/uploads/{rel}'
        product.updated_at = datetime.utcnow()
        db.commit()
        log_event(db, 'update', 'product_image', f'商品 {product.model} 重設圖片', str(product.id))
        return {'ok': True, 'image_url': f'/uploads/{rel}?v={int(datetime.utcnow().timestamp())}'}
    finally:
        db.close()

def create_or_update_quote(db, quote, form, product_ids, qtys):
    is_new = quote is None
    summary = quote_summary_from_request(db, form, product_ids, qtys)
    dealer = summary['dealer']
    if session.get('dealer_id') and not is_internal_logged_in():
        dealer = current_dealer(db)
        if dealer:
            summary['price_level'] = dealer.level or summary['price_level']
    parsed_date = datetime.strptime(form.get('quote_date', ''), '%Y-%m-%d').date() if form.get('quote_date') else date.today()
    quote_no = form.get('quote_no', '').strip()
    if not quote_no:
        quote_no = f"ZQ{parsed_date.strftime('%Y%m%d')}{db.query(models.Quote).count() + 1:03d}"

    if not quote:
        quote = models.Quote(quote_no=quote_no, quote_date=parsed_date, currency='NTD', customer_name=form.get('customer_name', '') or '未填', price_level=summary['price_level'])
        db.add(quote)
        db.flush()
    else:
        db.query(models.QuoteItem).filter(models.QuoteItem.quote_id == quote.id).delete()
        db.flush()

    quote.quote_no = quote_no
    quote.dealer_id = dealer.id if dealer else None
    quote.customer_name = form.get('customer_name', '')
    quote.contact_name = form.get('contact_name', '')
    quote.phone = form.get('phone', '')
    quote.email = form.get('email', '')
    quote.address = form.get('address', '')
    quote.attn = form.get('attn', '')
    quote.price_level = summary['price_level']
    quote.quote_date = parsed_date
    quote.note = summary['note']
    quote.product_subtotal = summary['product_subtotal']
    quote.planning_fee_total = summary['planning_fee_total']
    quote.setup_fee_total = summary['setup_fee_total']
    quote.dispatch_fee = summary['dispatch_fee']
    quote.planning_multiplier = summary['planning_multiplier']
    quote.setup_multiplier = summary['setup_multiplier']
    quote.dispatch_label = summary['dispatch_label']
    quote.lock_install_qty = summary['lock_install_qty']
    quote.lock_install_unit_price = summary['lock_install_unit_price']
    quote.lock_install_fee = summary['lock_install_fee']
    quote.curtain_type = summary.get('curtain_type','')
    quote.curtain_motor_price = summary.get('curtain_motor_price',0)
    quote.curtain_track_unit_price = summary.get('curtain_track_unit_price',0)
    quote.curtain_track_length = summary.get('curtain_track_length',0)
    quote.curtain_fabric_width = summary.get('curtain_fabric_width',0)
    quote.curtain_fabric_height = summary.get('curtain_fabric_height',0)
    quote.curtain_fabric_unit_price = summary.get('curtain_fabric_unit_price',0)
    quote.curtain_note = summary.get('curtain_note','')
    quote.curtain_rows_json = json.dumps(summary.get('curtain_rows', []), ensure_ascii=False)
    quote.curtain_install_qty = summary['curtain_install_qty']
    quote.curtain_install_unit = summary['curtain_install_unit']
    quote.curtain_install_amount = summary['curtain_install_amount']
    quote.weak_current_qty = summary['weak_current_qty']
    quote.weak_current_unit = summary['weak_current_unit']
    quote.weak_current_amount = summary['weak_current_amount']
    quote.hardware_qty = summary['hardware_qty']
    quote.hardware_unit = summary['hardware_unit']
    quote.hardware_amount = summary['hardware_amount']
    quote.water_elec_qty = summary['water_elec_qty']
    quote.water_elec_unit = summary['water_elec_unit']
    quote.water_elec_amount = summary['water_elec_amount']
    quote.custom_fee_json = json.dumps(summary.get('custom_fee_rows', []), ensure_ascii=False)
    quote.subtotal = summary['subtotal']
    quote.tax_amount = summary['tax_amount']
    quote.total_amount = summary['total_amount']
    quote.negotiated_total = summary.get('negotiated_total',0)
    quote.negotiated_discount_pct = summary.get('negotiated_discount_pct',0)
    quote.gross_profit_rate = summary['gross_profit_rate']
    quote.deposit_1 = summary['deposit_1']
    quote.deposit_2 = summary['deposit_2']
    quote.deposit_3 = summary['deposit_3']
    quote.payment_scheme = summary['payment_scheme']

    for idx, item in enumerate(summary['items'], start=1):
        product = item['product']
        db.add(models.QuoteItem(
            quote_id=quote.id,
            product_id=product.id,
            model=product.model,
            product_name=product.name,
            qty=item['qty'],
            unit=product.unit or '台',
            unit_price_twd=item['unit_price'],
            line_total_twd=item['line_total'],
            cost_total_twd=item['cost_total'],
            planning_fee_twd=item['planning_fee'],
            setup_fee_twd=item['setup_fee'],
            note=item['note'],
            image_path=get_image_src(product),
        ))
    db.commit()
    db.refresh(quote)
    log_event(db, 'create' if is_new else 'update', 'quote', f'報價單 {quote.quote_no} 已儲存', str(quote.id))
    return quote




@flask_app.route('/quotes/import_excel', methods=['POST'])
@require_internal_or_dealer_quote
def quote_import_excel():
    uploaded = request.files.get('quote_excel')
    if not uploaded or not uploaded.filename:
        return redirect('/quotes/new')
    try:
        wb = openpyxl.load_workbook(uploaded, data_only=True)
        ws = wb.active
        parsed = {
            'customer_name': ws['B5'].value or '',
            'contact_name': ws['B6'].value or '',
            'phone': ws['B7'].value or '',
            'email': ws['B8'].value or '',
            'address': ws['B9'].value or '',
            'quote_no': ws['B4'].value or '',
            'quote_date': (ws['H5'].value.strftime('%Y-%m-%d') if hasattr(ws['H5'].value, 'strftime') else str(ws['H5'].value or '')),
            'sales_name': ws['H6'].value or '',
            'sales_phone': ws['H7'].value or '',
            'sales_email': ws['H8'].value or '',
            'price_level': (ws['J6'].value or '市場報價'),
            'items': [],
        }
        for r in range(12, 45):
            item_no = ws.cell(r,1).value
            model = (ws.cell(r,2).value or '')
            name = (ws.cell(r,3).value or '')
            qty = ws.cell(r,4).value or 0
            unit = ws.cell(r,5).value or ''
            unit_price = ws.cell(r,6).value or 0
            line_total = ws.cell(r,7).value or 0
            note = ws.cell(r,8).value or ''
            if not name and not model:
                continue
            parsed['items'].append({'model': str(model).strip(), 'name': str(name).strip(), 'qty': int(safe_number(qty) or 0), 'unit': unit, 'unit_price': round(safe_number(unit_price),0), 'line_total': round(safe_number(line_total),0), 'note': str(note).strip()})
        session['imported_quote_payload'] = parsed
    except Exception as e:
        session['imported_quote_payload'] = {'error': str(e)}
    return redirect('/quotes/new?imported=1')

@flask_app.route('/quotes/new', methods=['GET', 'POST'])
@require_internal_or_dealer_quote
def quote_new_page():
    db = db_session()
    try:
        save_error = None
        if request.method == 'POST':
            if request.form.get('action_type') == 'preview':
                pass
            else:
                product_ids = request.form.getlist('product_ids')
                qtys = request.form.getlist('qtys')
                try:
                    quote = create_or_update_quote(db, None, request.form, product_ids, qtys)
                    return redirect(f'/quotes/{quote.id}')
                except Exception as e:
                    db.rollback()
                    save_error = f'建立報價單失敗：{e}'
        dealer_mode = bool(session.get('dealer_id') and not is_internal_logged_in())
        current_dealer_row = current_dealer(db) if dealer_mode else None
        dealers = [current_dealer_row] if dealer_mode and current_dealer_row else db.query(models.Dealer).order_by(models.Dealer.name).all()
        sales_people = db.query(models.SalesPerson).order_by(models.SalesPerson.name).all()
        products = db.query(models.Product).order_by(models.Product.category, models.Product.model).all()
        discount_map = get_price_setting_map(db)
        rates = get_rate_map(db)
        for p in products:
            p.usd_rate = rates['USD']
            p.rmb_rate = rates.get('RMB', 4.5)
        product_json = json.dumps([serialize_product(p, discount_map, rates['USD'], rates.get('RMB', 4.5)) for p in products], ensure_ascii=False)
        preview = None
        if request.method == 'POST' and request.form.get('action_type') == 'preview':
            preview = quote_summary_from_request(db, request.form, request.form.getlist('product_ids'), request.form.getlist('qtys'))
        initial_rows = []
        imported_payload = session.pop('imported_quote_payload', None) if request.args.get('imported') else None
        if imported_payload and not imported_payload.get('error'):
            form_defaults = imported_payload
            initial_rows = []
            for item in imported_payload.get('items', []):
                match = next((p for p in products if (item.get('model') and p.model == item.get('model')) or (item.get('name') and p.name == item.get('name'))), None)
                initial_rows.append({'product_id': match.id if match else '', 'qty': item.get('qty',0), 'note': item.get('note',''), 'search': item.get('model') or item.get('name')})
        else:
            form_defaults = {}
        existing_quotes = db.query(models.Quote).order_by(models.Quote.created_at.desc()).limit(100).all()
        return render_template('quote_form.html', dealers=dealers, sales_people=sales_people, products=products, quote=None, qty_map={}, note_text=DEFAULT_NOTE, price_levels=PRICE_LEVELS, product_json=product_json, preview=preview, dispatch_options=DISPATCH_OPTIONS, initial_rows=json.dumps(initial_rows, ensure_ascii=False), company_logo_src=get_company_logo_src(), existing_quotes=existing_quotes, dealer_mode=dealer_mode, dealer_fixed=current_dealer_row, initial_custom_fees='[]', initial_curtain_rows='[]', form_defaults=form_defaults, import_error=(imported_payload.get('error') if imported_payload else ''), save_error=save_error, today_str=date.today().isoformat())
    finally:
        db.close()


@flask_app.route('/quotes/<int:quote_id>')
def quote_detail(quote_id: int):
    db = db_session()
    try:
        quote = db.query(models.Quote).filter(models.Quote.id == quote_id).first()
        if not quote:
            abort(404)
        return render_template('quote_detail.html', quote=quote, output_name=quote_output_name(quote), display_items=build_display_items(quote), company_logo_src=get_company_logo_src())
    finally:
        db.close()


@flask_app.route('/quotes/<int:quote_id>/edit', methods=['GET', 'POST'])
@require_internal_or_dealer_quote
def quote_edit_page(quote_id: int):
    db = db_session()
    try:
        save_error = None
        quote = db.query(models.Quote).filter(models.Quote.id == quote_id).first()
        if not quote:
            abort(404)
        if request.method == 'POST' and request.form.get('action_type') != 'preview':
            try:
                quote = create_or_update_quote(db, quote, request.form, request.form.getlist('product_ids'), request.form.getlist('qtys'))
                return redirect(f'/quotes/{quote.id}')
            except Exception as e:
                db.rollback()
                save_error = f'更新報價單失敗：{e}'
        dealer_mode = bool(session.get('dealer_id') and not is_internal_logged_in())
        current_dealer_row = current_dealer(db) if dealer_mode else None
        dealers = [current_dealer_row] if dealer_mode and current_dealer_row else db.query(models.Dealer).order_by(models.Dealer.name).all()
        sales_people = db.query(models.SalesPerson).order_by(models.SalesPerson.name).all()
        products = db.query(models.Product).order_by(models.Product.category, models.Product.model).all()
        discount_map = get_price_setting_map(db)
        rates = get_rate_map(db)
        for p in products:
            p.usd_rate = rates['USD']
            p.rmb_rate = rates.get('RMB', 4.5)
        product_json = json.dumps([serialize_product(p, discount_map, rates['USD'], rates.get('RMB', 4.5)) for p in products], ensure_ascii=False)
        preview = None
        if request.method == 'POST' and request.form.get('action_type') == 'preview':
            preview = quote_summary_from_request(db, request.form, request.form.getlist('product_ids'), request.form.getlist('qtys'))
        initial_rows = [{"product_id": item.product_id, "qty": item.qty, "note": item.note or "", "search": ((item.product.model if item.product else "") or (item.product_name or ""))} for item in quote.items if item.product_id]
        existing_quotes = db.query(models.Quote).order_by(models.Quote.created_at.desc()).limit(100).all()
        return render_template('quote_form.html', dealers=dealers, sales_people=sales_people, products=products, quote=quote, qty_map={}, selected_ids=[], note_text=quote.note or DEFAULT_NOTE, price_levels=PRICE_LEVELS, product_json=product_json, preview=preview, dispatch_options=DISPATCH_OPTIONS, initial_rows=json.dumps(initial_rows, ensure_ascii=False), company_logo_src=get_company_logo_src(), existing_quotes=existing_quotes, dealer_mode=dealer_mode, dealer_fixed=current_dealer_row, initial_custom_fees=(quote.custom_fee_json or '[]'), initial_curtain_rows=(quote.curtain_rows_json or '[]'), form_defaults={}, import_error='', save_error=save_error, today_str=date.today().isoformat())
    finally:
        db.close()


@flask_app.route('/quotes/<int:quote_id>/excel')
def quote_excel_export(quote_id: int):
    db = db_session()
    try:
        quote = db.query(models.Quote).filter(models.Quote.id == quote_id).first()
        if not quote:
            abort(404)
        output_name = quote_output_name(quote)
        output_path = EXPORT_DIR / f'{output_name}.xlsx'
        export_quote_to_excel(quote, str(output_path), base_dir=str(BASE_DIR))
        log_event(db, 'export', 'quote_excel', f'匯出報價單 Excel：{quote.quote_no}', str(quote.id))
        return send_file(output_path, as_attachment=True, download_name=output_path.name)
    finally:
        db.close()


@flask_app.route('/quotes/<int:quote_id>/pdf')
def quote_pdf_export(quote_id: int):
    db = db_session()
    try:
        quote = db.query(models.Quote).filter(models.Quote.id == quote_id).first()
        if not quote:
            abort(404)
        output_name = quote_output_name(quote)
        output_path = EXPORT_DIR / f'{output_name}.pdf'
        export_quote_to_pdf(quote, str(output_path), base_dir=str(BASE_DIR))
        log_event(db, 'export', 'quote_pdf', f'匯出報價單 PDF：{quote.quote_no}', str(quote.id))
        return send_file(output_path, as_attachment=True, download_name=output_path.name)
    finally:
        db.close()






@flask_app.route('/admin/logs')
@require_internal_permission('all')
def admin_logs_page():
    db = db_session()
    try:
        rows = db.query(models.AuditLog).order_by(models.AuditLog.created_at.desc()).limit(500).all()
        return render_template('audit_logs.html', rows=rows)
    finally:
        db.close()

@flask_app.route('/admin/sales', methods=['GET', 'POST'])
@require_internal_permission('all')
def admin_sales_page():
    db = db_session()
    try:
        message = None
        if request.method == 'POST':
            name = (request.form.get('name') or '').strip()
            if name:
                row = db.query(models.SalesPerson).filter(models.SalesPerson.name == name).first()
                if not row:
                    row = models.SalesPerson(name=name)
                    db.add(row)
                row.phone = request.form.get('phone', '')
                row.email = request.form.get('email', '')
                row.updated_at = datetime.utcnow()
                db.commit()
                message = '業務資料已儲存。'
                log_event(db, 'update', 'sales', f'業務資料 {name} 已儲存')
        rows = db.query(models.SalesPerson).order_by(models.SalesPerson.name).all()
        return render_template('sales_admin.html', rows=rows, message=message)
    finally:
        db.close()





@flask_app.route('/admin/dealers', methods=['GET', 'POST'])
@require_internal_permission('all')
def admin_dealers_page():
    db = db_session()
    try:
        message = None
        if request.method == 'POST':
            action = request.form.get('action_type') or 'save'
            if action == 'create':
                name = (request.form.get('name') or '').strip()
                if name:
                    dealer = models.Dealer(name=name)
                    db.add(dealer)
                    db.flush()
                    dealer.level = (request.form.get('level') or '').strip()
                    dealer.access_key = normalize_password_input(request.form.get('access_key') or '0000') or '0000'
                    dealer.tax_id = normalize_tax_id_input(request.form.get('tax_id') or '')
                    dealer.address = (request.form.get('address') or '').strip()
                    dealer.phone = (request.form.get('phone') or '').strip()
                    dealer.sales_owner = (request.form.get('sales_owner') or '').strip()
                    dealer.payment_method = (request.form.get('payment_method') or '').strip()
                    dealer.closing_day = (request.form.get('closing_day') or '').strip()
                    dealer.payment_day = (request.form.get('payment_day') or '').strip()
                    dealer.note = (request.form.get('note') or '').strip()
                    dealer.signed_month = (request.form.get('signed_month') or '').strip()
                    dealer.can_view_products = 1 if request.form.get('can_view_products') else 0
                    dealer.can_export_prices = 1 if request.form.get('can_export_prices') else 0
                    dealer.can_create_quote = 1 if request.form.get('can_create_quote') else 0
                    db.commit()
                    log_event(db, 'create', 'dealer', f'新增經銷商：{dealer.name}', str(dealer.id))
                    message = '經銷商已新增。'
            else:
                dealer_id = int(request.form.get('dealer_id') or 0)
                dealer = db.query(models.Dealer).filter(models.Dealer.id == dealer_id).first()
                if dealer:
                    dealer.name = (request.form.get('name') or dealer.name or '').strip() or dealer.name
                    dealer.level = (request.form.get('level') or dealer.level or '').strip()
                    dealer.access_key = normalize_password_input(request.form.get('access_key') or dealer.access_key or '') or '0000'
                    dealer.tax_id = normalize_tax_id_input(request.form.get('tax_id') or '')
                    dealer.address = (request.form.get('address') or '').strip()
                    dealer.phone = (request.form.get('phone') or '').strip()
                    dealer.sales_owner = (request.form.get('sales_owner') or '').strip()
                    dealer.payment_method = (request.form.get('payment_method') or '').strip()
                    dealer.closing_day = (request.form.get('closing_day') or '').strip()
                    dealer.payment_day = (request.form.get('payment_day') or '').strip()
                    dealer.note = (request.form.get('note') or '').strip()
                    dealer.signed_month = (request.form.get('signed_month') or '').strip()
                    dealer.can_view_products = 1 if request.form.get('can_view_products') else 0
                    dealer.can_export_prices = 1 if request.form.get('can_export_prices') else 0
                    dealer.can_create_quote = 1 if request.form.get('can_create_quote') else 0
                    db.commit()
                    log_event(db, 'update', 'dealer', f'更新經銷商：{dealer.name}', str(dealer.id))
                    message = '經銷商資料已更新。'
        dealers = db.query(models.Dealer).order_by(models.Dealer.name).all()
        return render_template('dealer_admin.html', dealers=dealers, message=message)
    finally:
        db.close()

@flask_app.route('/admin/internal-users', methods=['GET', 'POST'])
@require_internal_permission('all')
def admin_internal_users_page():
    db = db_session()
    try:
        ensure_default_internal_users(db)
        message = None
        if request.method == 'POST':
            action_type = request.form.get('action_type', 'save')
            if action_type == 'create':
                username = (request.form.get('username') or '').strip()
                password = normalize_password_input(request.form.get('password') or '')
                display_name = (request.form.get('display_name') or '').strip()
                role = (request.form.get('role') or 'quote_only').strip()
                if username and password and len(password) <= 10:
                    exists = db.query(models.InternalUser).filter(models.InternalUser.username == username).first()
                    if exists:
                        message = f'帳號已存在：{username}'
                    else:
                        db.add(models.InternalUser(username=username, password=password, display_name=display_name or username, role=role, is_active=1, updated_at=datetime.utcnow()))
                        db.commit()
                        message = f'已新增內部帳號：{username}'
            elif action_type == 'delete':
                user_id = int(request.form.get('user_id') or 0)
                user = db.query(models.InternalUser).filter(models.InternalUser.id == user_id).first()
                if user:
                    if user.username == DEFAULT_SUPERVISOR_USERNAME:
                        message = '固定主管帳號不可刪除。'
                    elif user.id == session.get('internal_user_id'):
                        message = '目前登入中的帳號不可刪除。'
                    else:
                        deleted_name = user.username
                        db.delete(user)
                        db.commit()
                        message = f'已刪除內部帳號：{deleted_name}'
            else:
                user_id = int(request.form.get('user_id') or 0)
                user = db.query(models.InternalUser).filter(models.InternalUser.id == user_id).first()
                if user and user.username != DEFAULT_SUPERVISOR_USERNAME:
                    user.display_name = (request.form.get('display_name') or '').strip() or user.username
                    user.role = (request.form.get('role') or 'quote_only').strip()
                    new_password = normalize_password_input(request.form.get('password') or '')
                    if new_password:
                        user.password = new_password
                    user.is_active = 1 if request.form.get('is_active') else 0
                    user.updated_at = datetime.utcnow()
                    db.commit()
                    message = f'已更新內部帳號：{user.username}'
        users = db.query(models.InternalUser).order_by(models.InternalUser.role.desc(), models.InternalUser.username.asc()).all()
        return render_template('internal_users_admin.html', users=users, message=message, default_supervisor_username=DEFAULT_SUPERVISOR_USERNAME, current_internal_user_id=session.get('internal_user_id'))
    finally:
        db.close()


@flask_app.route('/admin/options', methods=['GET', 'POST'])
@require_internal_permission('all')
def admin_options_page():
    db = db_session()
    try:
        message = None
        if request.method == 'POST':
            opt_type = request.form.get('opt_type')
            name = (request.form.get('name') or '').strip()
            old_name = (request.form.get('old_name') or '').strip()
            model_cls = models.CategoryOption if opt_type == 'category' else models.StatusOption
            if name:
                row = db.query(model_cls).filter(model_cls.name == (old_name or name)).first()
                if row:
                    row.name = name
                else:
                    db.add(model_cls(name=name))
                db.commit()
                message = '選單資料已儲存。'
        categories, statuses = get_option_lists(db)
        return render_template('admin_options.html', categories=categories, statuses=statuses, message=message)
    finally:
        db.close()


@flask_app.route('/admin/price-files', methods=['GET', 'POST'])
@require_internal_permission('all')
def admin_price_files_page():
    db = db_session()
    try:
        message = None
        if request.method == 'POST':
            file = request.files.get('price_file')
            year = int(request.form.get('version_year') or datetime.now().year)
            effective_date = request.form.get('effective_date') or None
            note = request.form.get('note') or ''
            if file and file.filename:
                target = UPLOAD_DIR / 'price_files' / file.filename
                target.parent.mkdir(parents=True, exist_ok=True)
                file.save(target)
                db.add(models.PriceFileArchive(version_year=year, effective_date=datetime.strptime(effective_date, '%Y-%m-%d').date() if effective_date else None, file_name=file.filename, note=note))
                db.commit()
                message = '年度價格表已上傳。'
        rows = db.query(models.PriceFileArchive).order_by(models.PriceFileArchive.version_year.desc(), models.PriceFileArchive.created_at.desc()).all()
        return render_template('price_files.html', rows=rows, message=message)
    finally:
        db.close()


@flask_app.route('/admin/product-logs/<int:product_id>/delete/<int:log_id>', methods=['POST'])
@require_internal_permission('all')
def delete_product_log(product_id, log_id):
    password = request.form.get('password') or ''
    db = db_session()
    try:
        if normalize_password_input(password) != get_internal_password(db):
            flash('刪除失敗：密碼不正確')
            return redirect(f'/admin/products/{product_id}/edit')
        row = db.query(models.ProductChangeLog).filter(models.ProductChangeLog.id == log_id, models.ProductChangeLog.product_id == product_id).first()
        if row:
            db.delete(row)
            db.commit()
        return redirect(f'/admin/products/{product_id}/edit')
    finally:
        db.close()

if __name__ == '__main__':
    flask_app.run(host='127.0.0.1', port=8000, debug=True)
